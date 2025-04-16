import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def adjust_yunnan_table_format(data):
    """调整云南省表格格式"""
    try:
        # 检查是否是DataFrame
        if not isinstance(data, pd.DataFrame):
            return data
            
        # 重命名列（如果需要）
        column_mapping = {
            '用户类别': '电压等级',
            '基本电价': '基本电费',
            '电度电价': '电度电费'
        }
        data = data.rename(columns=lambda x: column_mapping.get(x, x))
        
        # 调整列顺序（根据云南省的要求）
        desired_columns = ['电压等级', '基本电费', '电度电费']
        existing_columns = [col for col in desired_columns if col in data.columns]
        other_columns = [col for col in data.columns if col not in desired_columns]
        data = data[existing_columns + other_columns]
        
        return data
    except Exception as e:
        print(f"调整云南省表格格式失败: {str(e)}")
        return data

def apply_yunnan_styles(worksheet):
    """应用云南省特有的样式"""
    try:
        # 设置标题样式
        title_font = Font(name='宋体', size=16, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        for row in range(1, 4):
            cell = worksheet.cell(row=row, column=1)
            cell.font = title_font
            cell.alignment = title_alignment
        
        # 设置表头样式
        header_font = Font(name='宋体', size=12, bold=True)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        for cell in worksheet[4]:  # 第4行是表头
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 设置数据行样式
        data_font = Font(name='宋体', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row in worksheet.iter_rows(min_row=5):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
        
        # 设置列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            adjusted_width = min(max_length + 4, 30)  # 限制最大宽度为30
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 设置行高
        worksheet.row_dimensions[1].height = 30  # 标题行
        worksheet.row_dimensions[2].height = 25  # 副标题行
        worksheet.row_dimensions[3].height = 25  # 单位行
        worksheet.row_dimensions[4].height = 40  # 表头行
        
        # 设置默认数据行高
        for i in range(5, worksheet.max_row + 1):
            worksheet.row_dimensions[i].height = 30
            
    except Exception as e:
        print(f"应用云南省样式失败: {str(e)}")

def merge_yunnan_empty_cells(worksheet, start_row, end_row, start_col, end_col, header_row=None):
    """云南省特有的单元格合并逻辑"""
    try:
        # 设置边框样式
        thin_side = Side(style='thin')
        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )
        
        # 获取特殊列的索引
        special_merge_columns = ['电压等级', '基本电费', '电度电费']
        special_col_indices = []
        
        if header_row:
            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet.cell(row=header_row, column=col)
                if header_cell.value in special_merge_columns:
                    special_col_indices.append(col)
        
        # 对特殊列应用合并规则
        for col in special_col_indices:
            current_value = None
            merge_start = None
            
            for row in range(start_row, end_row + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value == current_value and cell_value is not None:
                    continue
                elif merge_start and merge_start < row - 1:
                    # 合并前面的相同值单元格
                    worksheet.merge_cells(
                        start_row=merge_start,
                        end_row=row - 1,
                        start_column=col,
                        end_column=col
                    )
                    # 应用边框
                    for r in range(merge_start, row):
                        worksheet.cell(row=r, column=col).border = thin_border
                
                current_value = cell_value
                merge_start = row if cell_value is not None else None
            
            # 处理最后一组相同值
            if merge_start and merge_start < end_row:
                worksheet.merge_cells(
                    start_row=merge_start,
                    end_row=end_row,
                    start_column=col,
                    end_column=col
                )
                # 应用边框
                for r in range(merge_start, end_row + 1):
                    worksheet.cell(row=r, column=col).border = thin_border
                
    except Exception as e:
        print(f"云南省单元格合并失败: {str(e)}")

def add_yunnan_notes_with_title(worksheet, notes, start_row, max_col):
    """添加云南省特有的注释格式"""
    try:
        if not notes:
            return
            
        # 添加"注释内容"标题行
        title_row = start_row + 1
        title_cell = worksheet.cell(row=title_row, column=1, value="注释说明")
        
        # 合并标题行
        title_range = f'A{title_row}:{openpyxl.utils.get_column_letter(max_col)}{title_row}'
        worksheet.merge_cells(title_range)
        
        # 设置标题样式
        title_font = Font(name='宋体', size=12, bold=True)
        title_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        title_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 应用标题样式
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        
        # 设置注释样式
        note_font = Font(name='宋体', size=11)
        note_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 写入注释内容
        for i, note in enumerate(notes, 1):
            note_row = title_row + i
            note_cell = worksheet.cell(row=note_row, column=1, value=note)
            
            # 合并注释行
            note_range = f'A{note_row}:{openpyxl.utils.get_column_letter(max_col)}{note_row}'
            worksheet.merge_cells(note_range)
            
            # 应用注释样式
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # 设置行高
            worksheet.row_dimensions[note_row].height = 30
            
    except Exception as e:
        print(f"添加云南省注释失败: {str(e)}")

def apply_yunnan_cell_format(worksheet):
    """应用云南省特有的单元格格式"""
    try:
        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置对齐方式
        center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        # 设置字体
        regular_font = Font(name='宋体', size=11)
        
        # 应用格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = regular_font
        
        # 设置标题行特殊格式
        title_font = Font(name='宋体', size=16, bold=True)
        subtitle_font = Font(name='宋体', size=14, bold=True)
        
        for cell in worksheet[1]:
            cell.font = title_font
        for cell in worksheet[2:4]:
            for c in cell:
                c.font = subtitle_font
        
        # 设置表头行特殊格式
        header_font = Font(name='宋体', size=12, bold=True)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        for cell in worksheet[4]:
            cell.font = header_font
            cell.fill = header_fill
            
    except Exception as e:
        print(f"应用云南省单元格格式失败: {str(e)}") 