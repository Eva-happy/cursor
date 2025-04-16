import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import pdfplumber
import warnings
import re
import math
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import urllib3
import openpyxl
import io
import fitz  # PyMuPDF
import PIL.Image
import pyautogui
import subprocess
import win32com.client
import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
# 禁用SSL警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings('ignore')
def write_to_excel(all_pages_data, project_name=None, is_yunnan=False):
    """将数据写入Excel文件
    Args:
        all_pages_data: 所有页面的数据
        project_name: 项目名称
        is_yunnan: 是否是云南省的数据
    """
    try:
        if not all_pages_data:
            print("没有数据需要保存")
            return None
            
        print(f"共有 {len(all_pages_data)} 页数据需要处理")
        print("数据预览:")
        for i, page in enumerate(all_pages_data, 1):
            print(f"\n第 {i} 页:")
            print(f"标题: {page.get('title', '无')}")
            print(f"表格行数: {len(page.get('table', []))}")
            print(f"注释数: {len(page.get('notes', []))}")
        
        # 创建downloads目录（如果不存在）
        if not os.path.exists('downloads'):
            os.makedirs('downloads')
        
        # 生成基础文件名
        base_filename = project_name if project_name else f'电价表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        
        # 检查是否存在同名文件，并添加版本号
        version = 1
        while True:
            excel_file = os.path.join('downloads', f'{base_filename}V{version}.xlsx')
            if not os.path.exists(excel_file):
                break
            version += 1

        if is_yunnan:
            # 创建工作簿
            workbook = openpyxl.Workbook()

            # 处理每一页数据
            for page_index, page_data in enumerate(all_pages_data, 1):
                print(f"\n处理第 {page_index} 页...")
                sheet_name = f'第{page_index}页'
                
                # 创建工作表
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                else:
                    worksheet = workbook.create_sheet(title=sheet_name)
                
                current_row = 1
                
                # 云南省特殊处理
                # 根据页码应用不同的格式
                if page_index <= 3:  # 前三页
                    # 设置标题和其他信息
                    if 'title' in page_data and page_data['title']:
                        worksheet['A1'] = page_data['title']
                        worksheet.merge_cells(f'A1:J1')  # 合并到J列
                        current_row += 1
                        
                    if 'subtitle' in page_data and page_data['subtitle']:
                        worksheet[f'A{current_row}'] = page_data['subtitle']
                        worksheet.merge_cells(f'A{current_row}:J{current_row}')
                        current_row += 1
                        
                    if 'unit' in page_data and page_data['unit']:
                        worksheet[f'A{current_row}'] = page_data['unit']
                        worksheet.merge_cells(f'A{current_row}:J{current_row}')
                        current_row += 1
                        
                    # 写入文本数据
                    if 'table' in page_data:
                        for row_idx, row_data in enumerate(page_data['table'], current_row):
                            # 跳过空行和提示文本
                            if not row_data or any(text in str(row_data[0]) for text in ['程序会自动', '请确认']):
                                continue
                            for col_idx, cell_value in enumerate(row_data, 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.value = cell_value
                                # 设置左对齐，不自动换行
                                cell.alignment = openpyxl.styles.Alignment(
                                    horizontal='left',
                                    vertical='center',
                                    wrap_text=False
                                )
                else:  # 后三页
                    if 'table' in page_data:
                        # 写入文本数据
                        for row_idx, row_data in enumerate(page_data['table'], current_row):
                            # 跳过空行和提示文本
                            if not row_data or any(text in str(row_data[0]) for text in ['程序会自动', '请确认']):
                                continue
                            # 写入数据到第一列
                            cell = worksheet.cell(row=row_idx, column=1)
                            cell.value = row_data[0] if row_data else ""
                            
                            # 合并A到J列
                            worksheet.merge_cells(
                                start_row=row_idx,
                                end_row=row_idx,
                                start_column=1,
                                end_column=10  # J列
                            )
                            
                            # 设置居中对齐，自动换行（与前三行保持一致）
                            cell.alignment = openpyxl.styles.Alignment(
                                horizontal='left',
                                vertical='center',
                                wrap_text=True
                            )
                    
                    # 添加注释（如果有）
                    if 'notes' in page_data and page_data['notes']:
                        notes = page_data['notes']
                        start_row = worksheet.max_row + 2
                        add_yunnan_notes_with_title(worksheet, notes, start_row, 10)  # 10列
                
                # 设置列宽
                for col in range(1, 11):  # A到J列
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                
                # 合并空白单元格（仅前三页需要）
                if page_index <= 3:
                    merge_yunnan_empty_cells(worksheet, 5, worksheet.max_row, 1, worksheet.max_column, header_row=4)

            # 删除默认的Sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

        else:
            # 其他省份的处理逻辑
            # 创建工作簿
            workbook = openpyxl.Workbook()
            
            # 处理所有电价表页面（除最后一页）
            print("处理电价表页面...")
            for page_index in range(len(all_pages_data) - 1):
                data, notes, calc_df = process_first_page(all_pages_data[page_index])
                if data is not None:
                    # 为每个电价表页面创建不同的sheet名
                    sheet_name = f'电价表{page_index + 1}' if page_index > 0 else '电价表'
                    
                    # 创建工作表
                    worksheet = workbook.create_sheet(title=sheet_name)
                    
                    # 将DataFrame写入Excel
                    for r_idx, row in enumerate(data.values, 5):  # 从第5行开始写入数据
                        for c_idx, value in enumerate(row, 1):
                            worksheet.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 写入列名
                    for c_idx, column in enumerate(data.columns, 1):
                        worksheet.cell(row=4, column=c_idx, value=column)
                    
                    # 设置标题
                    title = all_pages_data[page_index].get('title', '')
                    subtitle = all_pages_data[page_index].get('subtitle', '')
                    unit = all_pages_data[page_index].get('unit', '')
                    
                    # 写入标题并合并单元格
                    max_col = len(data.columns)
                    worksheet['A1'] = title
                    worksheet.merge_cells(f'A1:{openpyxl.utils.get_column_letter(max_col)}1')
                    worksheet['A2'] = subtitle
                    worksheet.merge_cells(f'A2:{openpyxl.utils.get_column_letter(max_col)}2')
                    worksheet['A3'] = unit
                    worksheet.merge_cells(f'A3:{openpyxl.utils.get_column_letter(max_col)}3')
                    
                    # 合并空白单元格，标题行（第4行）不合并
                    header_row = 4
                    merge_empty_cells(worksheet, header_row + 1, worksheet.max_row, 1, worksheet.max_column, header_row=header_row)
                    
                    # 写入注释
                    if notes:
                        add_notes_with_title(worksheet, notes, worksheet.max_row + 2, max_col)
                    
                    # 应用单元格格式
                    apply_cell_format(worksheet)
            
            # 处理最后一页（计算说明）
            print("\n处理使用说明页面...")
            if len(all_pages_data) > 0:
                last_page = all_pages_data[-1]
                # 创建计算说明工作表
                calc_sheet = workbook.create_sheet(title='计算说明')
                
                # 设置标题
                title = last_page.get('title', '')
                subtitle = last_page.get('subtitle', '')
                unit = last_page.get('unit', '')
                
                calc_sheet['A1'] = title
                calc_sheet['A2'] = subtitle
                calc_sheet['A3'] = unit
                
                # 获取表格数据
                if 'table' in last_page:
                    table_data = last_page['table']
                    # 写入表格数据
                    for i, row in enumerate(table_data, start=4):
                        for j, value in enumerate(row, start=1):
                            calc_sheet.cell(row=i, column=j, value=value)
                    
                    # 合并标题行
                    max_col = max(len(row) for row in table_data)
                    for row in range(1, 4):
                        calc_sheet.merge_cells(f'A{row}:{openpyxl.utils.get_column_letter(max_col)}{row}')
                    
                    # 合并空白单元格，标题行（第4行）不合并
                    header_row = 4
                    merge_empty_cells(calc_sheet, header_row + 1, calc_sheet.max_row, 1, max_col, header_row=header_row)
                    
                    # 写入注释
                    if 'notes' in last_page:
                        notes = last_page['notes']
                        if notes:
                            add_notes_with_title(calc_sheet, notes, len(table_data) + 4, max_col)
                    
                    # 应用单元格格式
                    apply_cell_format(calc_sheet)
            
            # 删除默认的Sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
        
        # 保存Excel文件
        workbook.save(excel_file)
        
        print(f"Excel文件已保存到: {excel_file}")
        return excel_file
        
    except Exception as e:
        print(f"保存Excel文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
def add_notes_with_title(worksheet, notes, start_row, max_col):
    """添加注释标题和内容，并设置格式"""
    if not notes:
        return
        
    # 添加"注释内容"标题行
    title_row = start_row + 1
    title_cell = worksheet.cell(row=title_row, column=1, value="注释内容")
    
    # 合并标题行
    title_range = f'A{title_row}:{openpyxl.utils.get_column_letter(max_col)}{title_row}'
    worksheet.merge_cells(title_range)
    
    # 创建样式对象
    title_fill = openpyxl.styles.PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    title_font = openpyxl.styles.Font(color='FFFFFF', bold=True)
    title_alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 创建边框样式
    thin_side = openpyxl.styles.Side(style='thin')
    thin_border = openpyxl.styles.Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    
    # 为标题行的所有单元格应用边框和对齐方式
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=title_row, column=col)
        cell.border = thin_border
        cell.alignment = title_alignment
        if col == 1:
            cell.fill = title_fill
            cell.font = title_font
    
    # 设置标题行高
    worksheet.row_dimensions[title_row].height = 30
    
    # 写入注释内容
    note_alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for i, note in enumerate(notes, 1):
        note_row = title_row + i
        note_cell = worksheet.cell(row=note_row, column=1, value=note)
        
        # 合并注释行
        note_range = f'A{note_row}:{openpyxl.utils.get_column_letter(max_col)}{note_row}'
        worksheet.merge_cells(note_range)
        
        # 为注释行的所有单元格应用格式
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=note_row, column=col)
            cell.border = thin_border
            cell.alignment = note_alignment
        
        # 设置注释行高为38
        worksheet.row_dimensions[note_row].height = 38
def apply_cell_format(worksheet, header_row=4):
    """应用单元格格式：添加边框、对齐方式、自动调整列宽"""
    # 设置加粗边框和普通边框样式
    thick_side = openpyxl.styles.Side(style='medium')
    thin_side = openpyxl.styles.Side(style='thin')
    
    thick_border = openpyxl.styles.Border(
        left=thick_side,
        right=thick_side,
        top=thick_side,
        bottom=thick_side
    )
    
    thin_border = openpyxl.styles.Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    
    # 设置对齐方式
    center_alignment = openpyxl.styles.Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    left_alignment = openpyxl.styles.Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )
    
    # 获取工作表的有效范围
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    # 找到注释开始的行
    notes_start_row = None
    notes_title_row = None
    for row in range(max_row, 0, -1):
        cell = worksheet.cell(row=row, column=1)
        if cell.value == "注释内容":
            notes_title_row = row
            break
        elif cell.value and isinstance(cell.value, str) and ('注：' in cell.value or '备注：' in cell.value):
            notes_start_row = row
            break
    
    # 应用格式到所有单元格
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            
            # 判断是否是注释部分
            is_note_section = (notes_title_row and row >= notes_title_row) or (notes_start_row and row >= notes_start_row)
            
            # 设置边框
            if not is_note_section and row > header_row:
                cell.border = thick_border
            else:
                cell.border = thin_border
            
            # 设置对齐方式
            if is_note_section:
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
    
    # 自动调整列宽
    for col in range(1, max_col + 1):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col)
        
        # 获取该列所有单元格的内容长度
        for row in range(1, max_row + 1):
            cell = worksheet.cell(row=row, column=col)
            try:
                if cell.value:
                    # 计算内容的显示宽度（考虑换行符）
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        # 中文字符计为2个单位宽度
                        line_length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in line)
                        max_length = max(max_length, line_length)
            except:
                pass
        
        # 设置列宽（加上一些边距）
        adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
        worksheet.column_dimensions[column].width = adjusted_width
def setup_driver():
    """设置并返回WebDriver"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        driver_path = os.path.join(script_dir, "msedgedriver.exe")
        
        if not os.path.exists(driver_path):
            print(f"错误: msedgedriver.exe 不存在于 {driver_path}")
            return None
        
        edge_options = Options()
        # 基本设置
        edge_options.add_argument("--start-maximized")  # 最大化窗口
        edge_options.add_argument('--disable-gpu')  # 禁用GPU加速
        edge_options.add_argument('--no-sandbox')  # 禁用沙盒模式
        edge_options.add_argument('--disable-dev-shm-usage')  # 禁用共享内存
        edge_options.add_argument('--disable-extensions')  # 禁用扩展
        edge_options.add_argument('--disable-logging')  # 禁用日志
        edge_options.add_argument('--log-level=3')  # 最小化日志级别
        edge_options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
        
        # 创建自定义服务对象
        service = Service(
            driver_path,
            log_output=os.devnull  # 禁用服务日志输出
        )
        
        print("正在启动浏览器...")
        driver = webdriver.Edge(service=service, options=edge_options)
        
        # 设置超时时间
        driver.set_page_load_timeout(10)  # 页面加载超时时间
        driver.set_script_timeout(5)  # 脚本执行超时时间
        driver.implicitly_wait(3)  # 隐式等待时间
        
        print("浏览器启动完成")
        return driver
        
    except Exception as e:
        print(f"设置驱动程序失败: {str(e)}")
        return None
def wait_and_click_element(driver, element):
    """等待并点击元素"""
    try:
        # 确保元素在视图中
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)
        time.sleep(0.5)
        
        # 尝试多种点击方法
        try_methods = [
            lambda: element.click(),
            lambda: driver.execute_script("arguments[0].click();", element),
            lambda: ActionChains(driver).move_to_element(element).click().perform()
        ]
        
        for method in try_methods:
            try:
                method()
                return True
            except:
                continue
        
        return False
    except Exception as e:
        print(f"点击元素失败: {str(e)}")
        return False
def get_provinces(driver):
    """获取省份列表"""
    try:
        print("尝试点击切换地区按钮...")
        # 等待"切换地区"按钮可点击
        time.sleep(3)
        
        # 尝试多种定位方式找到切换地区按钮
        region_button_selectors = [
            "//span[contains(text(), '切换地区')]",
            "//span[@class='ant-dropdown-link']",
            "//div[contains(@class, 'cBtn')]/span",
            "//div[@data-v-869be0c2 and contains(@class, 'cBtn')]/span"
        ]
        
        region_button = None
        for selector in region_button_selectors:
            try:
                print(f"尝试使用选择器查找切换地区按钮: {selector}")
                elements = driver.find_elements(By.XPATH, selector)
                if elements:
                    region_button = elements[0]
                    print("找到切换地区按钮")
                    break
            except Exception as e:
                print(f"使用选择器 {selector} 查找切换地区按钮失败: {str(e)}")
                continue
        
        if not region_button:
            print("未找到切换地区按钮")
            return []
        
        # 点击切换地区按钮
        print("正在点击切换地区按钮...")
        if not wait_and_click_element(driver, region_button):
            print("点击切换地区按钮失败")
            # 尝试使用JavaScript点击
            try:
                driver.execute_script("arguments[0].click();", region_button)
                print("使用JavaScript成功点击切换地区按钮")
            except Exception as e:
                print(f"使用JavaScript点击切换地区按钮失败: {str(e)}")
                return []
        
        # 等待省份列表加载
        print("等待省份列表加载...")
        time.sleep(2)
        
        # 只获取省份元素（不包括城市）
        provinces = []
        expected_provinces = ['广东省', '广西壮族自治区', '云南省', '贵州省', '海南省']
        
        try:
            # 尝试直接通过cityMenu中的第一个ul来获取省份
            province_selector = "div.cityMenu > ul:first-child > li"
            province_elements = driver.find_elements(By.CSS_SELECTOR, province_selector)
            
            if not province_elements:
                print("未找到省份元素，尝试使用XPath选择器")
                province_elements = driver.find_elements(By.XPATH, "//div[contains(@class, 'cityMenu')]/ul[1]/li")
            
            if province_elements:
                print(f"找到 {len(province_elements)} 个省份元素")
                for element in province_elements:
                    try:
                        # 获取省份名称
                        h5_element = element.find_element(By.TAG_NAME, "h5")
                        province_name = h5_element.text.strip()
                        if province_name and province_name in expected_provinces:
                            provinces.append({"text": province_name, "element": element})
                            print(f"成功添加省份: {province_name}")
                    except Exception as e:
                        print(f"处理省份元素时出错: {str(e)}")
                        continue
            else:
                print("未找到省份元素")
                # 如果无法找到省份元素，尝试使用硬编码方式
                for province_name in expected_provinces:
                    try:
                        selector = f"//h5[text()='{province_name}']/parent::li"
                        element = driver.find_element(By.XPATH, selector)
                        provinces.append({"text": province_name, "element": element})
                        print(f"成功添加省份(硬编码): {province_name}")
                    except Exception as e:
                        print(f"查找省份 {province_name} 失败: {str(e)}")
                        continue
        except Exception as e:
            print(f"查找省份元素失败: {str(e)}")
        
        if not provinces:
            print("未找到任何省份，请检查网页结构是否已改变")
        else:
            print(f"共找到 {len(provinces)} 个省份")
        
        return provinces
    except Exception as e:
        print(f"获取省份列表失败: {str(e)}")
        return []
def get_cities(driver):
    """获取城市列表"""
    try:
        print("等待城市列表加载...")
        time.sleep(1)
        cities = []
        
        # 尝试使用多种选择器获取城市列表
        selectors = [
            "div.cityMenu > ul.cityList > li",
            "//div[contains(@class, 'cityMenu')]/ul[contains(@class, 'cityList')]/li",
            "//div[contains(@class, 'cityMenu')]/ul[2]/li"  # 第二个ul应该是城市列表
        ]
        
        city_elements = []
        for selector in selectors:
            try:
                print(f"尝试使用选择器: {selector}")
                if selector.startswith("//"):
                    elements = driver.find_elements(By.XPATH, selector)
                else:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                
                if elements:
                    city_elements = elements
                    print(f"使用选择器 {selector} 找到 {len(elements)} 个城市元素")
                    break
            except Exception as e:
                print(f"使用选择器 {selector} 查找城市失败: {str(e)}")
                continue
        
        if city_elements:
            for element in city_elements:
                try:
                    # 获取城市名称
                    h5_element = element.find_element(By.TAG_NAME, "h5")
                    city_name = h5_element.text.strip()
                    if city_name:
                        cities.append({"text": city_name, "element": element})
                        print(f"成功添加城市: {city_name}")
                except Exception as e:
                    print(f"处理城市元素时出错: {str(e)}")
                    continue
        else:
            print("未找到城市元素，尝试使用JavaScript方式获取")
            try:
                cities_js = driver.execute_script("""
                    var cities = [];
                    var elements = document.querySelectorAll('.cityMenu .cityList li h5');
                    for(var i=0; i<elements.length; i++) {
                        var el = elements[i];
                        if(el.textContent && el.textContent.trim()) {
                            cities.push({
                                text: el.textContent.trim(),
                                element: el.parentElement
                            });
                        }
                    }
                    return cities;
                """)
                
                if cities_js:
                    for city in cities_js:
                        cities.append({"text": city['text'], "element": city['element']})
                    print(f"通过JavaScript找到 {len(cities)} 个城市")
            except Exception as e:
                print(f"JavaScript获取城市失败: {str(e)}")
        
        if not cities:
            print("未找到任何城市，请检查网页结构是否已改变")
        else:
            print(f"共找到 {len(cities)} 个城市")
        
        return cities
    except Exception as e:
        print(f"获取城市列表失败: {str(e)}")
        return []
def get_page_info(driver):
    """获取页码信息"""
    try:
        print("等待资讯公告列表加载...")
        WebDriverWait(driver,2).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'infoList')]"))
        )
        time.sleep(1)
        
        total_text = driver.find_element(By.XPATH, "//li[@class='ant-pagination-total-text']").text
        total_count = int(total_text.split()[1])
        total_pages = (total_count + 4) // 5
        print(f"\n总共有 {total_pages} 页")
        
        return total_pages
    except Exception as e:
        print(f"获取页码信息失败: {str(e)}")
        return 0
def get_page_announcements(driver, page_number):
    """获取指定页面的公告列表"""
    try:
        if page_number > 1:
            page_button = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((By.XPATH, f"//li[contains(@class, 'ant-pagination-item') and @title='{page_number}']"))
            )
            wait_and_click_element(driver, page_button)
            time.sleep(1)
        
        announcements = []
        announcement_elements = WebDriverWait(driver, 2).until(
            EC.presence_of_all_elements_located((By.XPATH, "//div[@data-v-7707c85a and @class='list-item']"))
        )
        
        for element in announcement_elements:
            try:
                title_element = element.find_element(By.XPATH, ".//div[@data-v-7707c85a and @class='esp']")
                date_element = element.find_element(By.XPATH, ".//div[@data-v-7707c85a and @class='timeLine']")
                
                title = title_element.text.strip()
                date = date_element.text.strip()
                
                if title:
                    link_element = element.find_element(By.XPATH, ".//div[@data-v-7707c85a and @class='link']")
                    announcements.append({
                        "text": f"{title} ({date})",
                        "element": link_element,
                        "page": page_number
                    })
            except Exception as e:
                print(f"处理公告元素失败: {str(e)}")
                continue
        
        return announcements
    except Exception as e:
        print(f"获取第 {page_number} 页公告失败: {str(e)}")
        return []
def display_menu(items, title="请选择：", allow_return=True):
    """显示菜单并获取用户选择"""
    while True:
        if any('page' in item for item in items):
            pages = sorted(list(set(item['page'] for item in items if 'page' in item)))
            
            print("\n=== 可选页码 ===")
            for page in pages:
                print(f"{page}. 第 {page} 页")
            if allow_return:
                print("0. 返回上一步")
            
            try:
                page_choice = int(input("\n请输入页码: "))
                if page_choice == 0 and allow_return:
                    return None
                if page_choice not in pages:
                    print(f"无效的页码，请输入: {', '.join(map(str, pages))}")
                    continue
                
                page_items = [item for item in items if item.get('page') == page_choice]
                
                print(f"\n=== 第 {page_choice} 页的公告列表 ===")
                for i, item in enumerate(page_items, 1):
                    print(f"{i}. {item['text']}")
                if allow_return:
                    print("0. 返回上一步")
                
                try:
                    item_choice = int(input(f"\n请输入编号(0-{len(page_items)}): "))
                    if item_choice == 0 and allow_return:
                        continue
                    if 1 <= item_choice <= len(page_items):
                        selected_item = page_items[item_choice-1]
                        print(f"\n您选择了: {selected_item['text']}")
                        return selected_item
                    print(f"无效的选择，请输入0-{len(page_items)}之间的数字")
                except ValueError:
                    print("请输入有效的数字")
            except ValueError:
                print("请输入有效的页码")
        else:
            print(f"\n{title}")
            print("可选项：")
            for i, item in enumerate(items, 1):
                print(f"{i}. {item['text']}")
            if allow_return:
                print("0. 返回上一步")
            
            try:
                choice = int(input(f"\n请输入编号(0-{len(items)}): "))
                if choice == 0 and allow_return:
                    return None
                if 1 <= choice <= len(items):
                    selected_item = items[choice-1]
                    print(f"\n您选择了: {selected_item['text']}")
                    return selected_item
                print(f"无效的选择，请输入0-{len(items)}之间的数字")
            except ValueError:
                print("请输入有效的数字")
def check_and_process_content(driver, project_name=None, is_yunnan=False):
    """检查并处理页面内容"""
    try:
        if is_yunnan:
            # 云南省特殊处理
            print("等待页面完全加载...")
            time.sleep(3)  # 给页面更多加载时间
            
            # 直接使用PDF下载链接
            pdf_url = "https://95598.csg.cn/ucs/ma/wt/business/downloadFiles?documentId=bmZkd195eF90eWZ3L3Nkay9vc3NGaWxlL2RlYmM0MTZkZDVjMzRmYzFiYjMxOGZkMDlkYTNhY2ZlL1BERg==&documentType=pdf&documentName=downloadFile"
            print("使用PDF下载链接:", pdf_url)
            
            # 使用下载方法
            pdf_file = download_pdf(pdf_url, project_name)
            if pdf_file:
                print("PDF下载成功，开始处理内容...")
                all_pages_data = extract_table_from_pdf(pdf_file, is_yunnan=True)
                if all_pages_data:
                    print("成功提取PDF数据，开始生成Excel文件...")
                    excel_file = write_to_excel(all_pages_data, project_name, is_yunnan=True)
                    if excel_file:
                        print("\n开始处理图片复制粘贴...")
                        # 直接从网页复制图片到Excel
                        copy_image_from_pdf_to_excel(driver, excel_file, [1, 2, 3])
                    return excel_file
                else:
                    print("PDF数据提取失败")
                    return None
            else:
                print("PDF下载失败")
                return None
        else:
            # 其他省份的处理逻辑
            try:
                wait = WebDriverWait(driver, 2)

                # 首先尝试查找普通的PDF iframe
                try:
                    iframe = wait.until(
                        EC.presence_of_element_located((By.XPATH, "//iframe[contains(@src, 'downloadFiles')]"))
                    )
                    pdf_url = iframe.get_attribute('src')
                except TimeoutException:
                # 如果找不到普通的PDF iframe，尝试查找并处理带链接的iframe
                    print("等待iframe加载...")
                    iframe = wait.until(
                        EC.presence_of_element_located((By.TAG_NAME, "iframe"))
                    )
                    driver.switch_to.frame(iframe)
                
                    print("查找并点击PDF链接...")
                    link = wait.until(
                        EC.element_to_be_clickable((By.TAG_NAME, "a"))
                    )
                    pdf_url = link.get_attribute('href')
                    
                    # 使用JavaScript点击链接
                    driver.execute_script("arguments[0].click();", link)
                    
                    # 切回主文档
                    driver.switch_to.default_content()
                    
                    # 等待新页面加载
                    print("等待PDF页面加载...")
                    time.sleep(2)  # 给页面一些加载时间
                    
                    # 获取当前窗口句柄
                    current_window = driver.current_window_handle
                    
                    # 切换到新窗口（如果有）
                    all_windows = driver.window_handles
                    for window in all_windows:
                        if window != current_window:
                            driver.switch_to.window(window)
                            break
            
                # 处理PDF内容
                all_pages_data = process_pdf_content(driver, pdf_url, project_name)
                if all_pages_data:
                    # 写入Excel
                    excel_file = write_to_excel(all_pages_data, project_name, is_yunnan=False)
                    return excel_file
                return None
            
            except TimeoutException:
                # 处理在线预览内容
                page_data = process_preview_content(driver)
                if page_data:
                    # 将单页数据包装成列表
                    all_pages_data = [page_data]
                    # 写入Excel
                    excel_file = write_to_excel(all_pages_data, project_name, is_yunnan=False)
                    return excel_file
                return None
            
    except Exception as e:
        print(f"处理内容失败: {str(e)}")
        import traceback
        traceback.print_exc()  # 打印完整的错误堆栈
        return None
def process_preview_content(driver):
    """处理在线预览内容"""
    try:
        print("\n开始处理在线预览内容...")
        
        # 使用显式等待等待页面加载完成
        wait = WebDriverWait(driver, 2)
        
        # 等待标题元素加载
        print("等待页面加载...")
        title = ""
        try:
            # 尝试多种可能的标题选择器
            title_selectors = [
                "//h1",
                "//div[contains(@class, 'title')]",
                "//div[contains(@class, 'header')]/div",
                "//div[contains(text(), '深圳供电局有限公司')]"
            ]
            
            for selector in title_selectors:
                try:
                    title_element = wait.until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )
                    title = title_element.text.strip()
                    if title:
                        print(f"找到标题: {title}")
                        break
                except TimeoutException:
                    continue
                except Exception:
                    continue
            
            if not title:
                print("未找到标题，使用默认值")
                title = "深圳供电局有限公司电价信息"
        except Exception as e:
            print(f"获取标题失败: {str(e)}")
            title = "深圳供电局有限公司电价信息"
        
        # 获取执行时间
        print("获取执行时间...")
        execution_time = ""
        try:
            time_selectors = [
                "//*[contains(text(), '执行时间')]",
                "//div[contains(text(), '年') and contains(text(), '月')]",
                "//span[contains(text(), '执行时间')]"
            ]
            
            for selector in time_selectors:
                try:
                    time_element = wait.until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )
                    execution_time = time_element.text.strip()
                    if execution_time:
                        print(f"找到执行时间: {execution_time}")
                        break
                except TimeoutException:
                    continue
                except Exception:
                    continue
        except Exception as e:
            print(f"获取执行时间失败: {str(e)}")
        
        # 获取表格数据
        print("获取表格数据...")
        table_data = []
        try:
            # 等待表格加载
            table = wait.until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            
            # 等待表格内容加载完成
            wait.until(
                lambda d: len(d.find_elements(By.TAG_NAME, "tr")) > 0
            )
            
            # 获取所有表格
            tables = driver.find_elements(By.TAG_NAME, "table")
            print(f"找到 {len(tables)} 个表格")
            
            if tables:
                # 使用第一个表格
                main_table = tables[0]
                
                # 获取所有行（包括表头和数据行）
                rows = main_table.find_elements(By.TAG_NAME, "tr")
                print(f"找到 {len(rows)} 行数据")
                
                for row in rows:
                    # 尝试获取表头单元格
                    header_cells = row.find_elements(By.TAG_NAME, "th")
                    if header_cells:
                        row_data = [cell.text.strip() for cell in header_cells]
                    else:
                        # 获取数据单元格
                        cells = row.find_elements(By.TAG_NAME, "td")
                        row_data = [cell.text.strip() for cell in cells]
                    
                    if any(row_data):  # 只添加非空行
                        table_data.append(row_data)
                        print(f"添加行: {row_data}")
            
            if not table_data:
                print("未找到表格数据，尝试其他方法...")
                try:
                    # 等待表格主体加载
                    table_body = wait.until(
                        EC.presence_of_element_located((By.CLASS_NAME, "el-table__body-wrapper"))
                    )
                    
                    # 等待表头加载
                    table_header = wait.until(
                        EC.presence_of_element_located((By.CLASS_NAME, "el-table__header-wrapper"))
                    )
                    
                    # 获取表头
                    header_rows = table_header.find_elements(By.TAG_NAME, "tr")
                    for row in header_rows:
                        cells = row.find_elements(By.TAG_NAME, "th")
                        row_data = [cell.text.strip() for cell in cells]
                        if any(row_data):
                            table_data.append(row_data)
                            print(f"添加表头行: {row_data}")
                    
                    # 获取数据行
                    rows = table_body.find_elements(By.CLASS_NAME, "el-table__row")
                    for row in rows:
                        cells = row.find_elements(By.CLASS_NAME, "el-table__cell")
                        row_data = [cell.text.strip() for cell in cells]
                        if any(row_data):
                            table_data.append(row_data)
                            print(f"添加数据行: {row_data}")
                except TimeoutException:
                    print("等待表格元素超时")
                except Exception as e:
                    print(f"获取表格数据失败: {str(e)}")
        
        except TimeoutException:
            print("等待表格加载超时")
        except Exception as e:
            print(f"获取表格数据失败: {str(e)}")
        
        if not table_data:
            print("未能获取到任何表格数据")
            return None
        
        # 获取注释内容
        print("获取注释内容...")
        notes = []
        try:
            # 尝试多种可能的注释选择器
            note_selectors = [
                "//*[contains(text(), '注：')]",
                "//*[contains(text(), '注:')]",
                "//div[starts-with(text(), '注：')]",
                "//div[starts-with(text(), '注:')]",
                "//p[contains(text(), '注：')]",
                "//*[contains(text(), '备注：')]",
                "//*[contains(text(), '备注:')]"
            ]
            
            for selector in note_selectors:
                try:
                    note_elements = driver.find_elements(By.XPATH, selector)
                    for element in note_elements:
                        note_text = element.text.strip()
                        if note_text and note_text not in notes:
                            notes.append(note_text)
                            print(f"找到注释: {note_text}")
                except Exception:
                    continue
        except Exception as e:
            print(f"获取注释失败: {str(e)}")
        
        # 创建page_data字典
        print("创建数据字典...")
        page_data = {
            'title': title,
            'subtitle': execution_time,
            'table': table_data,
            'notes': notes
        }
        
        return page_data
        
    except Exception as e:
        print(f"处理在线预览内容失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
def process_first_page(page_data):
    """处理第一页数据（主电价表）"""
    try:
        if not page_data or 'table' not in page_data:
            print("没有找到表格数据")
            return None, None, None
            
        table_data = page_data['table']
        if not table_data:
            print("表格数据为空")
            return None, None, None
            
        # 提取注释
        notes = page_data.get('notes', [])
        
        # 直接将原始表格数据转换为DataFrame，不做任何处理
        df = pd.DataFrame(table_data)
        
        return df, notes, None
    except Exception as e:
        print(f"处理第一页数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None, None
def process_second_page(page_data):
    """处理第二页数据（使用说明）"""
    try:
        if not page_data or 'table' not in page_data:
            print("没有找到表格数据")
            return None
            
        table_data = page_data['table']
        if not table_data:
            print("表格数据为空")
            return None
            
        # 直接将原始表格数据转换为DataFrame，不做任何处理
        df = pd.DataFrame(table_data)
        
        return df
    except Exception as e:
        print(f"处理第二页数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
def process_pdf_content(driver, pdf_url, project_name=None):
    """处理PDF内容
    Args:
        driver: WebDriver对象
        pdf_url: PDF文件的URL
        project_name: 项目名称，用于文件命名
    """
    try:
        if not pdf_url:
            return None
            
        # 下载PDF文件
        pdf_file = download_pdf(pdf_url, project_name)
        if not pdf_file:
            return None
                
        # 检查PDF文件是否存在且大小正常
        if not os.path.exists(pdf_file) or os.path.getsize(pdf_file) == 0:
            return None
            
        # 提取表格数据
        all_pages_data = extract_table_from_pdf(pdf_file)
        if not all_pages_data:
            return None
            
        # 直接返回提取的数据，而不是Excel文件路径
        return all_pages_data
            
    except Exception as e:
        print(f"处理PDF内容失败: {str(e)}")
        return None
def download_pdf(pdf_url, project_name=None):
    """下载PDF文件
    Args:
        pdf_url: PDF文件的URL
        project_name: 项目名称，用于文件命名
    """
    try:
        # 创建下载目录
        if not os.path.exists('downloads'):
            os.makedirs('downloads')
        
        # 生成基础文件名
        base_filename = project_name if project_name else f'电价表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        
        # 检查是否存在同名文件，并添加版本号
        version = 1
        while True:
            pdf_file = os.path.join('downloads', f'{base_filename}V{version}.pdf')
            if not os.path.exists(pdf_file):
                break
            version += 1
        
        # 下载PDF
        print(f"正在下载PDF到: {pdf_file}")
        
        # 设置请求头
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'application/pdf,*/*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Referer': 'https://95598.csg.cn/'
        }
        
        # 禁用SSL验证警告
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # 使用session来处理请求
        session = requests.Session()
        session.verify = False  # 禁用SSL验证
        
        # 设置超时时间
        response = session.get(pdf_url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            # 检查内容类型和文件大小
            content_type = response.headers.get('content-type', '').lower()
            content_length = len(response.content)
            
            if content_length == 0:
                print("下载的内容为空")
                return None
            
            # 保存文件
            with open(pdf_file, 'wb') as f:
                f.write(response.content)
            
            # 验证保存的文件
            if os.path.exists(pdf_file) and os.path.getsize(pdf_file) > 0:
                if 'application/pdf' in content_type or response.content.startswith(b'%PDF'):
                    print("PDF下载完成")
                else:
                    print(f"警告：文件已保存，但内容类型不是PDF (Content-Type: {content_type})")
                return pdf_file
            else:
                print("保存的文件为空或不存在")
                return None
        else:
            print(f"下载失败，状态码: {response.status_code}")
            return None
    except requests.Timeout:
        print("下载超时")
        return None
    except requests.RequestException as e:
        print(f"下载请求失败: {str(e)}")
        return None
    except Exception as e:
        print(f"下载PDF时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
def extract_images_from_pdf_yunnan(pdf_path, page_num):
    """使用PyMuPDF从本地PDF提取图片,专门用于云南省"""
    try:
        doc = fitz.open(pdf_path)
        if 0 <= page_num - 1 < doc.page_count:
            page = doc[page_num - 1]
            image_list = page.get_images()
            
            if image_list:
                print(f"第 {page_num} 页发现 {len(image_list)} 张图片")
                result = []
                
                for img_index, img in enumerate(image_list, start=1):
                    try:
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        
                        if base_image:
                            # 直接使用原始图片数据
                            image_bytes = base_image["image"]
                            image_ext = base_image["ext"]
                            
                            # 创建一个临时的内存文件
                            img_stream = io.BytesIO()
                            # 写入图片数据
                            img_stream.write(image_bytes)
                            # 将指针移到开始位置
                            img_stream.seek(0)
                            
                            result.append({
                                'stream': img_stream,
                                'width': base_image.get("width", 800),
                                'height': base_image.get("height", 600),
                                'ext': image_ext
                            })
                            print(f"成功提取图片 {img_index}，格式：{image_ext}")
                    except Exception as e:
                        print(f"处理图片 {img_index} 时出错: {str(e)}")
                        continue
                
                return result
            else:
                print(f"第 {page_num} 页没有找到图片")
                return []
    except Exception as e:
        print(f"提取图片时出错: {str(e)}")
        return []
    finally:
        if 'doc' in locals():
            doc.close()
def extract_text_from_yunnan_pdf(pdf_path):
    """使用PyMuPDF从本地PDF提取文本和图片，专门用于云南省"""
    try:
        print(f"\n开始处理PDF文件: {pdf_path}")
        all_pages_data = []
        doc = fitz.open(pdf_path)
        total_pages = doc.page_count
        print(f"PDF共有 {total_pages} 页")
            
        for page_num in range(total_pages):
            try:
                print(f"\n处理第 {page_num + 1} 页...")
                page = doc[page_num]
                page_data = {}
                    
                # 提取图片（前三页）
                if page_num < 3:
                    print(f"正在提取第 {page_num + 1} 页的图片...")
                    image_data = extract_images_from_pdf_yunnan(pdf_path, page_num + 1)
                    if image_data:
                        print(f"成功提取到 {len(image_data)} 张图片")
                        page_data['images'] = image_data
                    else:
                        print("未找到图片")
                
                # 提取文本
                text = page.get_text()
                print(f"提取到的文本长度: {len(text)}")
                print(f"文本前100个字符预览: {text[:100]}")
                
                if not text:
                    print(f"第 {page_num + 1} 页文本为空")
                    if 'images' in page_data:
                        all_pages_data.append(page_data)
                    continue
                
                # 分行处理文本
                lines = text.split('\n')
                lines = [line.strip() for line in lines if line.strip()]
                print(f"处理后的文本行数: {len(lines)}")
                
                if not lines:
                    print(f"第 {page_num + 1} 页没有文本行")
                    if 'images' in page_data:
                        all_pages_data.append(page_data)
                    continue
                
                # 根据页码不同处理文本内容
                if page_num < 3:  # 前三页：保留表格结构
                    print("使用前三页的处理逻辑")
                    # 提取标题（第一行）
                    page_data['title'] = lines[0]
                    print(f"标题: {page_data['title']}")
                        
                    # 提取执行时间（如果有）
                    for line in lines[1:3]:
                        if '执行时间' in line:
                            page_data['subtitle'] = line
                            print(f"执行时间: {line}")
                            break
                    
                    # 使用正则表达式清理文本
                    cleaned_lines = []
                    current_line = []
                    
                    # 只处理第一页和第二页的表格数据
                    if page_num < 2:
                        for line in lines[1:]:  # 跳过标题
                            line = line.strip()
                            if line:
                                current_line.append(line)
                                if not re.search(r'[a-zA-Z0-9]$', line):  # 如果行不以字母或数字结尾
                                    if current_line:
                                        cleaned_lines.append([' '.join(current_line)])
                                        current_line = []
                        
                        if current_line:  # 添加最后一行
                            cleaned_lines.append([' '.join(current_line)])
                    
                    # 设置表格数据
                    if cleaned_lines:
                        page_data['table'] = cleaned_lines
                        print(f"表格数据行数: {len(cleaned_lines)}")
                    else:
                        page_data['table'] = []
                        print("表格数据为空")
                
                else:  # 后三页：按段落分割
                    print("使用后三页的处理逻辑")
                    paragraphs = []
                    current_paragraph = []
                        
                    # 直接从第一行开始处理，不跳过标题
                    for i, line in enumerate(lines):
                        line = line.strip()
                        if line:
                            # 如果是新段落的开始（数字序号开头）
                            if re.match(r'^\d+[\.\、]', line):
                                if current_paragraph:
                                    # 合并当前段落并添加
                                    paragraphs.append([' '.join(current_paragraph)])
                                current_paragraph = [line]
                            else:
                                current_paragraph.append(line)
                                # 检查是否是段落结束
                                next_line = lines[i + 1] if i + 1 < len(lines) else None
                                if (line.endswith('。') and 
                                    (not next_line or re.match(r'^\d+[\.\、]', next_line))):
                                    if current_paragraph:
                                        paragraphs.append([' '.join(current_paragraph)])
                                        current_paragraph = []
                        
                    if current_paragraph:  # 添加最后一个段落
                        paragraphs.append([' '.join(current_paragraph)])
                        
                    if paragraphs:
                        page_data['table'] = paragraphs
                        print(f"段落数: {len(paragraphs)}")
                    
                # 提取注释
                notes = []
                if page_num < 3:  # 只在前三页提取注释
                    for line in lines:
                        if any(line.startswith(prefix) for prefix in ['注：', '注:', '备注：', '备注:', '注释说明：']):
                            notes.append(line)
                    
                    if notes:
                        page_data['notes'] = notes
                        print(f"注释数: {len(notes)}")
                    else:
                        page_data['notes'] = []
                        print("注释数: 0")
                else:
                    page_data['notes'] = []
                    print("注释数: 0")
                    
                # 提取单位信息
                for line in lines:
                    if '单位' in line:
                        page_data['unit'] = line
                        print(f"单位信息: {line}")
                        break
                    
                all_pages_data.append(page_data)
                print(f"第 {page_num + 1} 页处理完成")
                    
            except Exception as e:
                print(f"处理第 {page_num + 1} 页时出错: {str(e)}")
                continue
        
        print(f"\n总共处理了 {len(all_pages_data)} 页数据")
        return all_pages_data
    
    except Exception as e:
        print(f"提取文本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        if 'doc' in locals():
            doc.close()
def extract_table_from_pdf(pdf_path, is_yunnan=False):
    """从PDF中提取表格数据、文本内容和图片"""
    try:
        if is_yunnan:
            # 云南省特殊处理
            return extract_text_from_yunnan_pdf (pdf_path)
        else:
            # 其他省份的处理逻辑
            try:
                all_pages_data = []
                with pdfplumber.open(pdf_path) as pdf:
                    for page_num, page in enumerate(pdf.pages, 1):
                        try:
                            print(f"\n处理第 {page_num} 页...")
                            page_data = {}
                            
                            # 提取文本
                            text = page.extract_text()
                            if not text:
                                print(f"第 {page_num} 页文本为空")
                                continue
                            
                            lines = text.split('\n')
                            if not lines:
                                print(f"第 {page_num} 页没有文本行")
                                continue
                            
                            # 提取标题
                            title_line = lines[0].strip() if lines else None
                            if title_line:
                                page_data['title'] = title_line
                                print(f"标题: {title_line}")
                            
                            # 提取执行时间
                            for line in lines[1:3]:
                                if '执行时间' in line:
                                    page_data['subtitle'] = line.strip()
                                    print(f"执行时间: {line.strip()}")
                                    break
                            
                            # 提取表格
                            tables = page.extract_tables()
                            if tables:
                                main_table = tables[0]  # 使用第一个表格
                                if main_table and len(main_table) > 0:
                                    print(f"表格列数: {len(main_table[0])}")
                                    print(f"表头内容: {main_table[0]}")
                                    
                                    # 清理表格数据（移除空行和空列）
                                    cleaned_table = []
                                    for row in main_table:
                                        if any(cell.strip() if isinstance(cell, str) else cell for cell in row):
                                            cleaned_row = [
                                                cell.strip() if isinstance(cell, str) else cell
                                                for cell in row
                                            ]
                                            cleaned_table.append(cleaned_row)
                                    
                                    if cleaned_table:
                                        page_data['table'] = cleaned_table
                                    else:
                                        print("表格数据清理后为空")
                                else:
                                    print("表格结构无效")
                            else:
                                print("未找到表格")
                            
                            # 提取注释
                            notes = []
                            current_note = ""
                            in_notes = False
                            
                            for line in lines:
                                line = line.strip()
                                # 增加对"备注："的识别
                                if line.startswith('注：') or line.startswith('注:') or line.startswith('注释说明：') or line.startswith('备注：') or line.startswith('备注:'):
                                    if current_note:
                                        notes.append(current_note.strip())
                                    current_note = line if not (line.startswith('注释说明：') or line == '备注：' or line == '备注:') else ""
                                    in_notes = True
                                elif in_notes and line and not line.startswith('执行时间'):
                                    try:
                                        if re.match(r'^\d+[\.\、]', line):
                                            if current_note:
                                                notes.append(current_note.strip())
                                            current_note = line
                                        else:
                                            current_note = (current_note + " " + line) if current_note else line
                                    except Exception as e:
                                        print(f"处理注释行时出错: {str(e)}")
                                        continue
                            
                            if current_note:
                                notes.append(current_note.strip())
                            
                            # 处理注释内容
                            processed_notes = []
                            if notes:
                                # 移除第一个注释的前缀（"注："或"备注："）
                                if notes[0].startswith(('注：', '注:', '备注：', '备注:')):
                                    for prefix in ['注：', '注:', '备注：', '备注:']:
                                        if notes[0].startswith(prefix):
                                            notes[0] = notes[0][len(prefix):].strip()
                                            break
                                processed_notes = [note.strip() for note in notes if note.strip()]
                            
                            if processed_notes:
                                page_data['notes'] = processed_notes
                                print(f"找到 {len(processed_notes)} 条注释")
                            
                            # 提取单位信息
                            for line in lines:
                                if '单位' in line:
                                    page_data['unit'] = line.strip()
                                    print(f"单位信息: {line.strip()}")
                                    break
                            
                            # 只有当页面包含必要的数据时才添加
                            if 'table' in page_data:
                                all_pages_data.append(page_data)
                                print(f"第 {page_num} 页处理完成")
                            else:
                                print(f"第 {page_num} 页缺少必要的表格数据，已跳过")
                            
                        except Exception as e:
                            print(f"处理第 {page_num} 页时出错: {str(e)}")
                            continue
                
                if not all_pages_data:
                    print("未能从PDF中提取到有效数据")
                    return None
                
                print(f"共处理了 {len(all_pages_data)} 页有效数据")
                return all_pages_data
                
            except Exception as e:
                print(f"提取表格失败: {str(e)}")
                import traceback
                traceback.print_exc()
                return None
    
    except Exception as e:
        print(f"提取表格失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
def merge_empty_cells(worksheet, start_row, end_row, start_col, end_col, header_row=None):
    """合并空白单元格与相邻的有内容单元格"""
    def get_merged_ranges():
        try:
            return list(worksheet.merged_cells.ranges)
        except Exception:
            return []
    
    def is_cell_in_merged_range(row, col, ranges):
        try:
            return any(r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col 
                      for r in ranges)
        except Exception:
            return False
    
    def try_merge_cells(start_row, end_row, start_col, end_col):
        try:
            # 验证合并范围的有效性
            if start_row > end_row or start_col > end_col:
                return False
            if start_row < 1 or start_col < 1:
                return False
            if end_row > worksheet.max_row or end_col > worksheet.max_column:
                return False
            
            # 检查是否与现有合并区域重叠
            for merged_range in get_merged_ranges():
                if (max(start_row, merged_range.min_row) <= min(end_row, merged_range.max_row) and
                    max(start_col, merged_range.min_col) <= min(end_col, merged_range.max_col)):
                    return False
            
            worksheet.merge_cells(
                start_row=start_row,
                end_row=end_row,
                start_column=start_col,
                end_column=end_col
            )
            return True
        except Exception:
            return False
    
    def is_empty_cell(cell):
        try:
            if cell is None:
                return True
            value = cell.value
            if value is None:
                return True
            if isinstance(value, str):
                return not value.strip()
            return False
        except Exception:
            return False
    
    try:
        merged_ranges = get_merged_ranges()
        
        # 第一步：优先处理第五行的横向合并，并检查是否可以向下合并到第七行
        if header_row and header_row + 1 <= end_row:
            fifth_row = header_row + 1
            col = start_col
            while col <= end_col:
                if is_cell_in_merged_range(fifth_row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= fifth_row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        col = merge_range.max_col + 1
                    continue
                
                current_cell = worksheet.cell(row=fifth_row, column=col)
                if not is_empty_cell(current_cell):
                    # 向右查找可以合并的空单元格，直到遇到非空单元格
                    merge_end_col = col
                    for next_col in range(col + 1, end_col + 1):
                        next_cell = worksheet.cell(row=fifth_row, column=next_col)
                        if (is_empty_cell(next_cell) and 
                            not is_cell_in_merged_range(fifth_row, next_col, merged_ranges)):
                            merge_end_col = next_col
                        else:
                            break
                    
                    if merge_end_col > col:
                        # 先进行横向合并
                        if try_merge_cells(fifth_row, fifth_row, col, merge_end_col):
                            merged_ranges = get_merged_ranges()
                            
                            # 检查第六行和第七行的对应列是否为空
                            can_merge_down = True
                            max_merge_row = min(fifth_row + 2, end_row)  # 最多合并到第七行
                            
                            # 检查是否可以向下合并
                            for check_row in range(fifth_row + 1, max_merge_row + 1):
                                for check_col in range(col, merge_end_col + 1):
                                    check_cell = worksheet.cell(row=check_row, column=check_col)
                                    if not is_empty_cell(check_cell) or is_cell_in_merged_range(check_row, check_col, merged_ranges):
                                        can_merge_down = False
                                        break
                                if not can_merge_down:
                                    break
                            
                            # 如果可以向下合并，则合并到第七行（或最大可能行）
                            if can_merge_down and max_merge_row > fifth_row:
                                # 先取消之前的横向合并
                                worksheet.unmerge_cells(
                                    start_row=fifth_row,
                                    end_row=fifth_row,
                                    start_column=col,
                                    end_column=merge_end_col
                                )
                                # 进行完整的合并（横向+纵向）
                                if try_merge_cells(fifth_row, max_merge_row, col, merge_end_col):
                                    merged_ranges = get_merged_ranges()
                    
                    col = merge_end_col + 1
                else:
                    col += 1
        
        # 第二步：处理A-C列的横向合并（跳过第五行）
        text_col_end = min(3, end_col)  # A-C列（如果存在）
        for row in range(start_row, end_row + 1):
            if header_row and row == header_row + 1:  # 跳过第五行
                continue
            
            col = start_col
            while col <= text_col_end:
                if is_cell_in_merged_range(row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        col = merge_range.max_col + 1
                    continue
                
                current_cell = worksheet.cell(row=row, column=col)
                if not is_empty_cell(current_cell):
                    # 向右查找可以合并的空单元格
                    merge_end_col = col
                    for next_col in range(col + 1, text_col_end + 1):
                        next_cell = worksheet.cell(row=row, column=next_col)
                        if (is_empty_cell(next_cell) and 
                            not is_cell_in_merged_range(row, next_col, merged_ranges)):
                            merge_end_col = next_col
                        else:
                            break
                    
                    if merge_end_col > col:
                        if try_merge_cells(row, row, col, merge_end_col):
                            merged_ranges = get_merged_ranges()
                    col = merge_end_col + 1
                else:
                    col += 1
        
        # 第三步：处理A-C列的纵向合并（包括第五行）
        for col in range(1, text_col_end + 1):  # 仅处理A-C列
            row = start_row
            while row <= end_row:
                if is_cell_in_merged_range(row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        row = merge_range.max_row + 1
                    continue
                
                current_cell = worksheet.cell(row=row, column=col)
                if not is_empty_cell(current_cell):
                    # 获取当前单元格所在的横向合并范围
                    current_merge = next((r for r in merged_ranges 
                                        if r.min_row == row and r.min_col <= col <= r.max_col), None)
                    
                    merge_start_col = current_merge.min_col if current_merge else col
                    merge_end_col = current_merge.max_col if current_merge else col
                    
                    # 向下查找可以合并的空行
                    merge_end_row = row
                    for next_row in range(row + 1, end_row + 1):
                        # 检查下一行的整个范围是否为空
                        is_range_empty = True
                        for check_col in range(merge_start_col, merge_end_col + 1):
                            next_cell = worksheet.cell(row=next_row, column=check_col)
                            if not is_empty_cell(next_cell) or is_cell_in_merged_range(next_row, check_col, merged_ranges):
                                is_range_empty = False
                                break
                        
                        if is_range_empty:
                            merge_end_row = next_row
                        else:
                            break
                    
                    if merge_end_row > row:
                        if current_merge:
                            # 先取消横向合并
                            worksheet.unmerge_cells(
                                start_row=row,
                                end_row=row,
                                start_column=merge_start_col,
                                end_column=merge_end_col
                            )
                        # 进行完整的合并
                        if try_merge_cells(row, merge_end_row, merge_start_col, merge_end_col):
                            merged_ranges = get_merged_ranges()
                    row = merge_end_row + 1
                else:
                    row += 1
        
        # 第四步：处理其他列的纵向合并（D列及之后的列）
        for col in range(text_col_end + 1, end_col + 1):
            row = start_row
            while row <= end_row:
                if is_cell_in_merged_range(row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        row = merge_range.max_row + 1
                    continue
                
                current_cell = worksheet.cell(row=row, column=col)
                if not is_empty_cell(current_cell):
                    # 向下查找可以合并的空单元格
                    merge_end_row = row
                    for next_row in range(row + 1, end_row + 1):
                        next_cell = worksheet.cell(row=next_row, column=col)
                        if (is_empty_cell(next_cell) and 
                            not is_cell_in_merged_range(next_row, col, merged_ranges)):
                            merge_end_row = next_row
                        else:
                            break
                    
                    if merge_end_row > row:
                        if try_merge_cells(row, merge_end_row, col, col):
                            merged_ranges = get_merged_ranges()
                    row = merge_end_row + 1
                else:
                    row += 1
                    
    except Exception as e:
        print(f"合并单元格时出错: {str(e)}")
        # 发生错误时不影响程序继续运行
def get_mouse_position():
    """
    获取鼠标位置的辅助函数
    使用方法：运行此函数，然后将鼠标移动到目标位置，按Ctrl+C结束
    """
    try:
        print("请将鼠标移动到目标位置...")
        print("按Ctrl+C停止并显示坐标")
        while True:
            x, y = pyautogui.position()
            position_str = f'当前鼠标位置: X: {x} Y: {y}'
            print(position_str, end='\r')
            time.sleep(0.1)
    except KeyboardInterrupt:
        print("\n已记录最后位置")
        return x, y
def copy_image_from_pdf_to_excel(driver, excel_path, page_numbers):
    """
    使用自动化方式将网页中的图片复制到Excel
    Args:
        driver: WebDriver对象
        excel_path: Excel文件路径
        page_numbers: 需要处理的页码列表
    """
    excel = None
    workbook = None
    max_retries = 3  # 最大重试次数
    
    try:
        # 转换为绝对路径
        excel_abs_path = os.path.abspath(excel_path)
        print(f"Excel路径: {excel_abs_path}")
        
        # 检查文件是否存在
        if not os.path.exists(excel_abs_path):
            print(f"Excel文件不存在: {excel_abs_path}")
            return
        
        # 创建临时文件夹（如果不存在）
        temp_dir = os.path.join('downloads', 'temp')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            print(f"创建临时目录: {temp_dir}")
        
        # 打开Excel文件（带重试）
        for attempt in range(max_retries):
            try:
                print(f"尝试打开Excel... (尝试 {attempt + 1}/{max_retries})")
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                excel.DisplayAlerts = False
                workbook = excel.Workbooks.Open(excel_abs_path)
                print("Excel打开成功")
                break
            except Exception as e:
                print(f"打开Excel失败 (尝试 {attempt + 1}): {str(e)}")
                if attempt == max_retries - 1:
                    raise
                if excel:
                    try:
                        excel.Quit()
                    except:
                        pass
                time.sleep(2)  # 等待一段时间后重试
        
        # 设置pyautogui安全设置
        pyautogui.FAILSAFE = True
        pyautogui.PAUSE = 1.5
        
        for page_num in page_numbers:
            try:
                print(f"\n处理第 {page_num} 页...")
                
                # 询问是否需要重新获取坐标
                response = input(f"\n是否需要重新获取第 {page_num} 页的图片坐标？(y/n): ")
                if response.lower() == 'y':
                    print("\n请获取图片左上角坐标:")
                    start_x, start_y = get_mouse_position()
                    print(f"\n左上角坐标: X: {start_x} Y: {start_y}")
                    
                    print("\n请获取图片右下角坐标:")
                    end_x, end_y = get_mouse_position()
                    print(f"\n右下角坐标: X: {end_x} Y: {end_y}")
                else:
                    # 使用上一次的坐标
                    if 'start_x' not in locals():
                        print("没有保存的坐标，请至少获取一次坐标")
                        continue
                
                print("\n请切换到PDF网页，3秒后将自动截图...")
                time.sleep(3)
                
                # 计算区域大小
                width = end_x - start_x
                height = end_y - start_y
                
                # 保存截图到临时文件（带重试）
                temp_image = os.path.join(temp_dir, f'temp_image_{page_num}.png')
                screenshot_success = False
                
                for attempt in range(max_retries):
                    try:
                        print(f"尝试截图... (尝试 {attempt + 1}/{max_retries})")
                        screenshot = pyautogui.screenshot(region=(start_x, start_y, width, height))
                        screenshot.save(temp_image)
                        
                        if os.path.exists(temp_image) and os.path.getsize(temp_image) > 0:
                            screenshot_success = True
                            print("截图成功")
                            break
                    except Exception as e:
                        print(f"截图失败 (尝试 {attempt + 1}): {str(e)}")
                        if attempt == max_retries - 1:
                            print("截图失败，跳过当前页面")
                            continue
                        time.sleep(2)
                
                if not screenshot_success:
                    continue
                
                # 插入图片到Excel（带重试）
                for attempt in range(max_retries):
                    try:
                        print(f"尝试插入图片... (尝试 {attempt + 1}/{max_retries})")
                        
                        # 激活工作表
                        worksheet = workbook.Worksheets(f'第{page_num}页')
                        worksheet.Activate()
                        
                        # 选择插入位置
                        paste_row = 6
                        cell = worksheet.Cells(paste_row, 2)
                        cell.Select()
                        
                        # 获取插入位置
                        left = cell.Left
                        top = cell.Top
                        
                        # 插入图片
                        shape = worksheet.Shapes.AddPicture(
                            os.path.abspath(temp_image),
                            LinkToFile=False,
                            SaveWithDocument=True,
                            Left=left,
                            Top=top,
                            Width=600,
                            Height=400
                        )
                        
                        # 调整行高
                        worksheet.Rows(paste_row).RowHeight = 400
                        
                        print(f"第 {page_num} 页图片插入成功")
                        break
                        
                    except Exception as e:
                        print(f"插入图片失败 (尝试 {attempt + 1}): {str(e)}")
                        if attempt == max_retries - 1:
                            print("插入图片失败，跳过当前页面")
                        time.sleep(2)
                
            except Exception as e:
                print(f"处理第 {page_num} 页时出错: {str(e)}")
                continue
        
        # 保存Excel文件（带重试）
        for attempt in range(max_retries):
            try:
                print(f"\n尝试保存Excel... (尝试 {attempt + 1}/{max_retries})")
                workbook.Save()
                print(f"Excel文件已保存到: {excel_abs_path}")
                break
            except Exception as e:
                print(f"保存Excel失败 (尝试 {attempt + 1}): {str(e)}")
                if attempt == max_retries - 1:
                    print("保存Excel失败")
                time.sleep(2)
        
        # 询问是否关闭Excel
        response = input("\n所有图片处理完成，是否关闭Excel？(y/n): ")
        if response.lower() == 'y':
            try:
                print("关闭Excel...")
                if workbook:
                    workbook.Close()
                if excel:
                    excel.Quit()
            except Exception as e:
                print(f"关闭Excel时出错: {str(e)}")
        else:
            print("Excel保持打开状态")
        
    except Exception as e:
        print(f"处理过程出错: {str(e)}")
        import traceback
        traceback.print_exc()
        
    finally:
        # 清理临时文件
        try:
            if os.path.exists(temp_dir):
                import shutil
                shutil.rmtree(temp_dir)
                print("临时目录已清理")
        except Exception as e:
            print(f"清理临时目录失败: {str(e)}")
        
        # 确保Excel正确关闭
        if 'response' in locals() and response.lower() == 'y':
            try:
                if workbook:
                    workbook.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
            except:
                pass
def get_yunnan_price_items(driver):
    """获取云南省电价项目列表"""
    try:
        print("等待电价项目列表加载...")
        price_items = []
        
        # 等待项目列表加载
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'list-item')]"))
        )
        
        # 获取所有项目元素
        items = driver.find_elements(By.XPATH, "//div[contains(@class, 'list-item')]")
        
        for item in items:
            try:
                # 获取项目标题
                title = item.find_element(By.XPATH, ".//div[contains(@class, 'esp')]").text.strip()
                # 获取日期
                date = item.find_element(By.XPATH, ".//div[contains(@class, 'timeLine')]").text.strip()
                
                if title and '电价' in title:  # 只添加包含"电价"的项目
                    link_element = item.find_element(By.XPATH, ".//div[contains(@class, 'link')]")
                    price_items.append({
                        "text": f"{title} ({date})",
                        "element": link_element
                    })
                    print(f"找到电价项目: {title}")
            except Exception as e:
                print(f"处理项目元素时出错: {str(e)}")
                continue
        
        return price_items
    except Exception as e:
        print(f"获取电价项目列表失败: {str(e)}")
        return []

def adjust_yunnan_table_format(data):
    """调整云南省表格格式"""
    try:
        # 检查是否是DataFrame
        if not isinstance(data, pd.DataFrame):
            return data
            
        # 重命名列（如果需要）
        column_mapping = {
            '用户类别': '电压等级',
            '基本电价': '基本电费',
            '电度电价': '电度电费'
        }
        data = data.rename(columns=lambda x: column_mapping.get(x, x))
        
        # 调整列顺序（根据云南省的要求）
        desired_columns = ['电压等级', '基本电费', '电度电费']
        existing_columns = [col for col in desired_columns if col in data.columns]
        other_columns = [col for col in data.columns if col not in desired_columns]
        data = data[existing_columns + other_columns]
        
        return data
    except Exception as e:
        print(f"调整云南省表格格式失败: {str(e)}")
        return data
def apply_province_specific_format(worksheet, is_yunnan=False):
    """根据省份应用特定的格式"""
    if is_yunnan:
        apply_yunnan_styles(worksheet)
        apply_yunnan_cell_format(worksheet)
    else:
        apply_cell_format(worksheet)  # 通用格式
def apply_yunnan_styles(worksheet):
    """应用云南省特有的样式"""
    try:
        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置对齐方式
        center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        left_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )
        
        # 设置字体
        regular_font = Font(name='宋体', size=11)
        title_font = Font(name='宋体', size=16, bold=True)
        subtitle_font = Font(name='宋体', size=14, bold=True)
        
        # 获取工作表名称
        sheet_name = worksheet.title
        page_number = int(sheet_name.replace('第', '').replace('页', ''))
        
        # 应用基础格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.font = regular_font
                cell.alignment = left_alignment  # 默认使用左对齐
        
        if page_number <= 3:  # 前三页
            # 获取所有合并单元格的范围
            merged_ranges = worksheet.merged_cells.ranges
            
            # 对合并的单元格应用居中对齐
            for merged_range in merged_ranges:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = center_alignment
            
            # 设置标题行格式（第1行）
            for cell in worksheet[1]:
                cell.font = title_font
            
            # 设置副标题行格式（第2行）
            for cell in worksheet[2]:
                cell.font = title_font  # 使用与第1行相同的字体
            
            # 设置单位行格式（第3行）
            for cell in worksheet[3]:
                cell.font = subtitle_font
            
        else:  # 后三页
            # 前三行使用左对齐、自动换行
            for row in range(1, 4):
                for cell in worksheet[row]:
                    cell.alignment = left_alignment
                    if row == 1:
                        cell.font = title_font
                    elif row == 2:
                        cell.font = title_font
                    else:  # row == 3
                        cell.font = subtitle_font
                
                # 合并单元格（A到J列）
                if not any(worksheet.merged_cells.ranges, key=lambda r: r.min_row == row):
                    worksheet.merge_cells(f'A{row}:J{row}')
            
            # 其余行使用与前三行一致的格式
            for row in range(4, worksheet.max_row + 1):
                for cell in worksheet[row]:
                    cell.alignment = left_alignment
        
        # 设置行高
        worksheet.row_dimensions[1].height = 35  # 标题行
        worksheet.row_dimensions[2].height = 35  # 副标题行（与第1行相同）
        worksheet.row_dimensions[3].height = 25  # 单位行
        
        if page_number <= 3:  # 前三页
            worksheet.row_dimensions[4].height = 30  # 普通行高
            worksheet.row_dimensions[5].height = 30
        else:  # 后三页
            worksheet.row_dimensions[i].height = 47  # 修改为47
            
    except Exception as e:
        print(f"应用云南省样式失败: {str(e)}")
        import traceback
        traceback.print_exc()

def merge_yunnan_empty_cells(worksheet, start_row, end_row, start_col, end_col, header_row=None):
    """云南省特有的单元格合并逻辑"""
    try:
        # 设置边框样式
        thin_side = Side(style='thin')
        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )
        
        # 获取特殊列的索引
        special_merge_columns = ['电压等级', '基本电费', '电度电费']
        special_col_indices = []
        
        if header_row:
            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet.cell(row=header_row, column=col)
                if header_cell.value in special_merge_columns:
                    special_col_indices.append(col)
        
        # 对特殊列应用合并规则
        for col in special_col_indices:
            current_value = None
            merge_start = None
            
            for row in range(start_row, end_row + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value == current_value and cell_value is not None:
                    continue
                elif merge_start and merge_start < row - 1:
                    # 合并前面的相同值单元格
                    worksheet.merge_cells(
                        start_row=merge_start,
                        end_row=row - 1,
                        start_column=col,
                        end_column=col
                    )
                    # 应用边框
                    for r in range(merge_start, row):
                        worksheet.cell(row=r, column=col).border = thin_border
                
                current_value = cell_value
                merge_start = row if cell_value is not None else None
            
            # 处理最后一组相同值
            if merge_start and merge_start < end_row:
                worksheet.merge_cells(
                    start_row=merge_start,
                    end_row=end_row,
                    start_column=col,
                    end_column=col
                )
                # 应用边框
                for r in range(merge_start, end_row + 1):
                    worksheet.cell(row=r, column=col).border = thin_border
                
    except Exception as e:
        print(f"云南省单元格合并失败: {str(e)}")

def add_yunnan_notes_with_title(worksheet, notes, start_row, max_col):
    """添加云南省特有的注释格式"""
    try:
        if not notes:
            return
            
        # 添加"注释内容"标题行
        title_row = start_row + 1
        title_cell = worksheet.cell(row=title_row, column=1, value="注释说明")
        
        # 合并标题行
        title_range = f'A{title_row}:{openpyxl.utils.get_column_letter(max_col)}{title_row}'
        worksheet.merge_cells(title_range)
        
        # 设置标题样式
        title_font = Font(name='宋体', size=12, bold=True)
        title_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        title_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 应用标题样式
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        
        # 设置注释样式
        note_font = Font(name='宋体', size=11)
        note_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 写入注释内容
        for i, note in enumerate(notes, 1):
            note_row = title_row + i
            note_cell = worksheet.cell(row=note_row, column=1, value=note)
            
            # 合并注释行
            note_range = f'A{note_row}:{openpyxl.utils.get_column_letter(max_col)}{note_row}'
            worksheet.merge_cells(note_range)
            
            # 应用注释样式
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # 设置行高
            worksheet.row_dimensions[note_row].height = 30
            
    except Exception as e:
        print(f"添加云南省注释失败: {str(e)}")

def apply_yunnan_cell_format(worksheet):
    """应用云南省特有的单元格格式"""
    try:
        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置对齐方式
        center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        # 设置字体
        regular_font = Font(name='宋体', size=11)
        
        # 应用格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = regular_font
        
        # 设置标题行特殊格式
        title_font = Font(name='宋体', size=16, bold=True)
        subtitle_font = Font(name='宋体', size=14, bold=True)
        
        for cell in worksheet[1]:
            cell.font = title_font
        for cell in worksheet[2:4]:
            for c in cell:
                c.font = subtitle_font
        
        # 设置表头行特殊格式
        header_font = Font(name='宋体', size=12, bold=True)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        for cell in worksheet[4]:
            cell.font = header_font
            cell.fill = header_fill
            
    except Exception as e:
        print(f"应用云南省单元格格式失败: {str(e)}")
def main():
    """主函数"""
    print("=== 南方电网电价信息提取程序 ===")
    print("正在初始化程序...")
    
    driver = None
    max_retries = 2
    
    try:
        driver = setup_driver()
        if not driver:
            return
        
        while True:  # 主循环
            retry_count = 0
            success = False
            
            while retry_count < max_retries and not success:
                try:
                    print(f"\n正在加载网页... (尝试 {retry_count + 1}/{max_retries})")
                    driver.get("https://95598.csg.cn/#/gd/serviceInquire/information/list")
                    
                    WebDriverWait(driver, 1).until(
                        EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'ant-dropdown-link')]"))
                    )
                    
                    success = True
                    print("网页加载完成")
                    
                except Exception as e:
                    print(f"加载失败: {str(e)}")
                    retry_count += 1
                    if retry_count < max_retries:
                        print("正在重试...")
                        time.sleep(2)
                    else:
                        print("加载失败，请检查网络连接后重试")
                        return
            
            # 获取省份列表
            provinces = get_provinces(driver)
            if not provinces:
                print("未找到省份列表")
                return
            
            # 选择省份
            selected_province = display_menu(provinces, "请选择省份：")
            if selected_province is None:
                continue
            if not wait_and_click_element(driver, selected_province['element']):
                print("点击省份失败")
                continue
            
            # 获取城市列表
            cities = get_cities(driver)
            if not cities:
                print("未找到城市列表")
                continue
            
            # 选择城市
            selected_city = display_menu(cities, "请选择城市：")
            if selected_city is None:
                continue
            if not wait_and_click_element(driver, selected_city['element']):
                print("点击城市失败")
                continue
            
            # 云南省特殊处理
            if selected_province['text'] == '云南省':
                try:
                    print("\n检测到云南省，执行特殊处理流程...")
                    
                    # 点击营商环境
                    print("等待营商环境按钮加载...")
                    business_env = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//li[@role='menuitem']//span[text()='营商环境']"))
                    )
                    
                    if not wait_and_click_element(driver, business_env):
                        print("点击营商环境失败")
                        continue
                    
                    # 点击信息公开
                    print("等待信息公开按钮加载...")
                    info_disclosure = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'container')]//span[contains(text(), '信息公开')]"))
                    )
                    
                    if not wait_and_click_element(driver, info_disclosure):
                        print("点击信息公开失败")
                        continue
                    
                    # 点击电价及收费标准
                    print("等待电价及收费标准按钮加载...")
                    price_standard = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'list-item')]//div[contains(text(), '电价及收费标准')]"))
                    )
                    
                    if not wait_and_click_element(driver, price_standard):
                        print("点击电价及收费标准失败")
                        continue
                    
                    time.sleep(2)  # 等待页面加载
                    
                    # 获取电价项目列表
                    price_items = get_yunnan_price_items(driver)
                    if not price_items:
                        print("未找到电价项目列表")
                        continue
                    
                    # 选择项目
                    selected_item = display_menu(price_items, "请选择电价项目：")
                    if selected_item is None:
                        continue
                    
                    print(f"\n正在处理: {selected_item['text']}")
                    if not wait_and_click_element(driver, selected_item['element']):
                        print("点击项目失败")
                        continue
                    
                    time.sleep(2)  # 等待页面加载
                    
                    # 检查是否有PDF并处理
                    excel_file = check_and_process_content(driver, selected_item['text'], is_yunnan=True)
                    if excel_file:
                        print(f"\n处理完成，数据已保存到: {excel_file}")
                    else:
                        print("\n处理失败")
                
                except Exception as e:
                    print(f"云南省特殊处理失败: {str(e)}")
                    continue
            else:
                # 其他省份的原有处理逻辑# 获取总页数
                total_pages = get_page_info(driver)
                if total_pages == 0:
                    print("获取页码信息失败")
                    continue
                
                while True:  # 页码选择循环
                    print("\n=== 可选页码 ===")
                    for page in range(1, total_pages + 1):
                        print(f"{page}. 第 {page} 页")
                    print("0. 返回上一步")
                    
                    try:
                        page_choice = int(input("\n请输入页码: "))
                        if page_choice == 0:
                            break
                        
                        if 1 <= page_choice <= total_pages:
                            announcements = get_page_announcements(driver, page_choice)
                            if not announcements:
                                print(f"第 {page_choice} 页没有找到公告")
                                continue
                            
                            print(f"\n=== 第 {page_choice} 页的公告列表 ===")
                            for i, item in enumerate(announcements, 1):
                                print(f"{i}. {item['text']}")
                            print("0. 返回上一步")
                            
                            try:
                                item_choice = int(input(f"\n请输入编号(0-{len(announcements)}): "))
                                if item_choice == 0:
                                    continue
                                
                                if 1 <= item_choice <= len(announcements):
                                    selected_announcement = announcements[item_choice-1]
                                    print(f"\n您选择了: {selected_announcement['text']}")
                                    
                                    print(f"\n正在处理: {selected_announcement['text']}")
                                    if not wait_and_click_element(driver, selected_announcement['element']):
                                        print("点击公告失败")
                                        continue
                                    
                                    time.sleep(2)  # 等待页面加载
                                    
                                    # 检查是否有PDF并处理
                                    excel_file = check_and_process_content(driver, selected_announcement['text'], is_yunnan=False)
                                    if excel_file:
                                        print(f"\n处理完成，数据已保存到: {excel_file}")
                                    else:
                                        print("\n处理失败")
                                    
                                    while True:  # 操作选择循环
                                        print("\n请选择操作：")
                                        print("1. 继续浏览")
                                        print("2. 返回上一步")
                                        print("3. 退出程序")
                                        
                                        try:
                                            choice = int(input("\n请输入编号(1-3): "))
                                            if choice == 1:
                                                break
                                            elif choice == 2:
                                                break
                                            elif choice == 3:
                                                return
                                            else:
                                                print("无效的选择，请输入1-3之间的数字")
                                        except ValueError:
                                            print("请输入有效的数字")
                                else:
                                    print(f"无效的选择，请输入0-{len(announcements)}之间的数字")
                            except ValueError:
                                print("请输入有效的数字")
                        else:
                            print(f"无效的页码，请输入0-{total_pages}之间的数字")
                    except ValueError:
                        print("请输入有效的页码")
            
            # 操作选择循环
            while True:
                print("\n请选择操作：")
                print("1. 继续浏览")
                print("2. 返回上一步")
                print("3. 退出程序")
                
                try:
                    choice = int(input("\n请输入编号(1-3): "))
                    if choice == 1:
                        break
                    elif choice == 2:
                        break
                    elif choice == 3:
                        return
                    else:
                        print("无效的选择，请输入1-3之间的数字")
                except ValueError:
                    print("请输入有效的数字")
        
    except Exception as e:
        print(f"\n程序运行出错: {str(e)}")
    finally:
        if driver:
            driver.quit()
if __name__ == "__main__":
    main()
