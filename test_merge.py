import openpyxl
import re

def merge_empty_cells_test(worksheet, start_row, end_row, start_col, end_col, header_row=None):
    """测试版本的合并空白单元格函数"""
    def get_merged_ranges():
        return list(worksheet.merged_cells.ranges)
    
    def is_cell_in_merged_range(row, col, ranges):
        return any(r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col 
                  for r in ranges)
    
    def try_merge_cells(start_row, end_row, start_col, end_col):
        try:
            worksheet.merge_cells(
                start_row=start_row,
                end_row=end_row,
                start_column=start_col,
                end_column=end_col
            )
            return True
        except ValueError:
            return False
    
    def is_empty_cell(cell):
        if cell is None:
            return True
        value = cell.value
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        return False
    
    def find_data_start_row():
        def has_standalone_number(row_num):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row_num, column=col)
                if not is_empty_cell(cell):
                    value = str(cell.value).strip()
                    if re.match(r'^-?\d+\.?\d*$', value):
                        return True
            return False

        for row in range(end_row, start_row - 1, -1):
            if has_standalone_number(row):
                data_start = row
                while row > start_row:
                    row -= 1
                    if not has_standalone_number(row):
                        return row + 1
                return start_row
        return None

    # 找到数据开始行
    data_start_row = find_data_start_row()
    if not data_start_row:
        return
    
    merged_ranges = get_merged_ranges()
    
    # 处理表头部分（先纵向后横向）
    # 表头纵向合并
    for col in range(start_col, end_col + 1):
        for row in range(data_start_row - 1, start_row - 1, -1):
            if is_cell_in_merged_range(row, col, merged_ranges):
                continue
            
            current_cell = worksheet.cell(row=row, column=col)
            if not is_empty_cell(current_cell):
                # 向下合并
                merge_end = row
                for next_row in range(row + 1, data_start_row):
                    next_cell = worksheet.cell(row=next_row, column=col)
                    if not is_empty_cell(next_cell) or is_cell_in_merged_range(next_row, col, merged_ranges):
                        break
                    merge_end = next_row
                
                if merge_end > row:
                    try_merge_cells(row, merge_end, col, col)
                    merged_ranges = get_merged_ranges()
            elif is_empty_cell(current_cell):
                # 向上合并
                prev_row = row - 1
                while prev_row >= start_row:
                    prev_cell = worksheet.cell(row=prev_row, column=col)
                    if not is_empty_cell(prev_cell) and not is_cell_in_merged_range(prev_row, col, merged_ranges):
                        try_merge_cells(prev_row, row, col, col)
                        merged_ranges = get_merged_ranges()
                        break
                    prev_row -= 1
    
    # 表头横向合并
    for row in range(data_start_row - 1, start_row - 1, -1):
        col = start_col
        while col <= end_col:
            if is_cell_in_merged_range(row, col, merged_ranges):
                merge_range = next((r for r in merged_ranges 
                                  if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                if merge_range:
                    col = merge_range.max_col + 1
                continue
            
            current_cell = worksheet.cell(row=row, column=col)
            if current_cell.value and not is_empty_cell(current_cell):
                merge_end = col
                for next_col in range(col + 1, end_col + 1):
                    next_cell = worksheet.cell(row=row, column=next_col)
                    if (is_empty_cell(next_cell) and 
                        not is_cell_in_merged_range(row, next_col, merged_ranges)):
                        merge_end = next_col
                    else:
                        break
                
                if merge_end > col:
                    try_merge_cells(row, row, col, merge_end)
                    merged_ranges = get_merged_ranges()
                col = merge_end + 1
            else:
                col += 1
    
    # 处理数据部分（先纵向后横向）
    # 数据纵向合并
    for col in range(start_col, end_col + 1):
        row = data_start_row
        while row <= end_row:
            if is_cell_in_merged_range(row, col, merged_ranges):
                merge_range = next((r for r in merged_ranges 
                                  if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                if merge_range:
                    row = merge_range.max_row + 1
                continue
            
            current_cell = worksheet.cell(row=row, column=col)
            if current_cell.value and not is_empty_cell(current_cell):
                merge_end = row
                for next_row in range(row + 1, end_row + 1):
                    next_cell = worksheet.cell(row=next_row, column=col)
                    if (is_empty_cell(next_cell) and 
                        not is_cell_in_merged_range(next_row, col, merged_ranges)):
                        merge_end = next_row
                    else:
                        break
                
                if merge_end > row:
                    try_merge_cells(row, merge_end, col, col)
                    merged_ranges = get_merged_ranges()
                row = merge_end + 1
            else:
                row += 1
    
    # 数据横向合并
    for row in range(data_start_row, end_row + 1):
        col = start_col
        while col <= end_col:
            if is_cell_in_merged_range(row, col, merged_ranges):
                merge_range = next((r for r in merged_ranges 
                                  if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                if merge_range:
                    col = merge_range.max_col + 1
                continue
            
            current_cell = worksheet.cell(row=row, column=col)
            if current_cell.value and not is_empty_cell(current_cell):
                merge_end = col
                for next_col in range(col + 1, end_col + 1):
                    next_cell = worksheet.cell(row=row, column=next_col)
                    if (is_empty_cell(next_cell) and 
                        not is_cell_in_merged_range(row, next_col, merged_ranges)):
                        merge_end = next_col
                    else:
                        break
                
                if merge_end > col:
                    try_merge_cells(row, row, col, merge_end)
                    merged_ranges = get_merged_ranges()
                col = merge_end + 1
            else:
                col += 1

# 创建测试工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active

# 测试数据
test_data = [
    ['用电分类', '电压等级', '基本电价', ''],
    ['', '', '（元/千瓦·月）', '（元/千伏安·月）'],
    ['工商业用电', '1-10kV', '45.00', ''],
    ['工商业用电', '35-110kV', '42.50', ''],
    ['工商业用电', '220kV及以上', '40.00', ''],
]

# 写入测试数据
for i, row in enumerate(test_data, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=value)

# 应用合并逻辑
merge_empty_cells_test(ws, 1, len(test_data), 1, len(test_data[0]), header_row=1)

# 保存测试结果
wb.save('test_merge_result.xlsx')
print("测试完成，结果已保存到 test_merge_result.xlsx") 