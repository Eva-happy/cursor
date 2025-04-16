import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_loan_template():
    wb = openpyxl.Workbook()
    
    # 创建参数表
    params = wb.active
    params.title = "参数设置"
    setup_params_sheet(params)
    
    # 创建等额本息表
    equal_payment = wb.create_sheet("等额本息")
    setup_equal_payment_sheet(equal_payment)
    
    # 创建等额本金表
    equal_principal = wb.create_sheet("等额本金")
    setup_equal_principal_sheet(equal_principal)
    
    # 创建先息后本表
    interest_first = wb.create_sheet("先息后本")
    setup_interest_first_sheet(interest_first)
    
    # 创建动态还款表
    dynamic = wb.create_sheet("动态还款")
    setup_dynamic_sheet(dynamic)
    
    wb.save('loan_calculation_template.xlsx')

def setup_params_sheet(ws):
    # 设置标题
    ws['A1'] = '贷款计算基础参数'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # 设置参数
    params = [
        ('贷款本金', 1000000),
        ('年利率', 0.045),
        ('贷款年限', 5),
        ('还款频率', '年'),
    ]
    
    for i, (param, value) in enumerate(params, start=2):
        ws[f'A{i}'] = param
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # 设置现金流表
    ws['E2'] = '年份'
    ws['F2'] = '可用现金流'
    ws['E2'].font = Font(bold=True)
    ws['F2'].font = Font(bold=True)
    
    for i in range(5):
        ws[f'E{i+3}'] = f'第{i+1}年'
        ws[f'F{i+3}'] = 300000  # 示例现金流

def setup_equal_payment_sheet(ws):
    # 设置标题和公式说明
    ws['A1'] = '等额本息还款计划'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # 设置表头
    headers = ['期数', '期初余额', '还款额', '本金', '利息', '期末余额', '累计还款']
    for i, header in enumerate(headers, start=1):
        ws[f'{get_column_letter(i)}3'] = header
        ws[f'{get_column_letter(i)}3'].font = Font(bold=True)
    
    # 设置公式
    for i in range(1, 6):  # 假设5年
        row = i + 3
        ws[f'A{row}'] = i  # 期数
        if i == 1:
            ws[f'B{row}'] = '=参数设置!B2'  # 期初余额等于贷款本金
        else:
            ws[f'B{row}'] = f'=G{row-1}'  # 期初余额等于上期期末余额
        
        # PMT公式计算还款额
        ws[f'C{row}'] = '=PMT(参数设置!B3,参数设置!B4,-参数设置!B2)'
        ws[f'D{row}'] = f'=B{row}*参数设置!B3'  # 本金
        ws[f'E{row}'] = f'=C{row}-D{row}'  # 利息
        ws[f'F{row}'] = f'=B{row}-D{row}'  # 期末余额
        ws[f'G{row}'] = f'=IF(ROW()=4,C4,G{row-1}+C{row})'  # 累计还款

def setup_equal_principal_sheet(ws):
    # 设置标题和公式说明
    ws['A1'] = '等额本金还款计划'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # 设置表头
    headers = ['期数', '期初余额', '还款额', '本金', '利息', '期末余额', '累计还款']
    for i, header in enumerate(headers, start=1):
        ws[f'{get_column_letter(i)}3'] = header
        ws[f'{get_column_letter(i)}3'].font = Font(bold=True)
    
    # 设置公式
    for i in range(1, 6):  # 假设5年
        row = i + 3
        ws[f'A{row}'] = i  # 期数
        if i == 1:
            ws[f'B{row}'] = '=参数设置!B2'  # 期初余额等于贷款本金
        else:
            ws[f'B{row}'] = f'=F{row-1}'  # 期初余额等于上期期末余额
        
        ws[f'D{row}'] = '=参数设置!B2/参数设置!B4'  # 每期本金金额
        ws[f'E{row}'] = f'=B{row}*参数设置!B3'  # 利息
        ws[f'C{row}'] = f'=D{row}+E{row}'  # 还款额
        ws[f'F{row}'] = f'=B{row}-D{row}'  # 期末余额
        ws[f'G{row}'] = f'=IF(ROW()=4,C4,G{row-1}+C{row})'  # 累计还款

def setup_interest_first_sheet(ws):
    # 设置标题和公式说明
    ws['A1'] = '先息后本还款计划'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # 设置表头
    headers = ['期数', '期初余额', '还款额', '本金', '利息', '期末余额', '累计还款']
    for i, header in enumerate(headers, start=1):
        ws[f'{get_column_letter(i)}3'] = header
        ws[f'{get_column_letter(i)}3'].font = Font(bold=True)
    
    # 设置公式
    for i in range(1, 6):  # 假设5年
        row = i + 3
        ws[f'A{row}'] = i  # 期数
        if i == 1:
            ws[f'B{row}'] = '=参数设置!B2'  # 期初余额等于贷款本金
        else:
            ws[f'B{row}'] = f'=F{row-1}'  # 期初余额等于上期期末余额
        
        ws[f'E{row}'] = f'=B{row}*参数设置!B3'  # 利息
        ws[f'D{row}'] = f'=IF(A{row}=参数设置!B4,B{row},0)'  # 本金（最后一期还清）
        ws[f'C{row}'] = f'=D{row}+E{row}'  # 还款额
        ws[f'F{row}'] = f'=B{row}-D{row}'  # 期末余额
        ws[f'G{row}'] = f'=IF(ROW()=4,C4,G{row-1}+C{row})'  # 累计还款

def setup_dynamic_sheet(ws):
    # 设置标题和公式说明
    ws['A1'] = '动态还款计划'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:I1')
    
    # 设置表头
    headers = ['期数', '期初余额', '可用现金流', '还款额', '本金', '利息', '期末余额', '累计还款', '现金流结余']
    for i, header in enumerate(headers, start=1):
        ws[f'{get_column_letter(i)}3'] = header
        ws[f'{get_column_letter(i)}3'].font = Font(bold=True)
    
    # 设置公式
    for i in range(1, 6):  # 假设5年
        row = i + 3
        ws[f'A{row}'] = i  # 期数
        if i == 1:
            ws[f'B{row}'] = '=参数设置!B2'  # 期初余额等于贷款本金
        else:
            ws[f'B{row}'] = f'=G{row-1}'  # 期初余额等于上期期末余额
        
        ws[f'C{row}'] = f'=参数设置!F{i+2}'  # 可用现金流
        ws[f'F{row}'] = f'=B{row}*参数设置!B3'  # 利息
        ws[f'E{row}'] = f'=MIN(B{row},MAX(0,C{row}-F{row}))'  # 本金还款
        ws[f'D{row}'] = f'=E{row}+F{row}'  # 还款额
        ws[f'G{row}'] = f'=B{row}-E{row}'  # 期末余额
        ws[f'H{row}'] = f'=IF(ROW()=4,D4,H{row-1}+D{row})'  # 累计还款
        ws[f'I{row}'] = f'=C{row}-D{row}'  # 现金流结余

if __name__ == "__main__":
    create_loan_template() 