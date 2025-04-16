
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from math import ceil
# 创建工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = "增强版成本模板V1"
# ---------------------------
# ① 黄色输入区：飞机架数（支持1-999架）
# ---------------------------
ws["A1"] = "飞机架数（1-999）："
ws["B1"] = 15  # 默认值示例
# 添加数据验证，限定输入范围1~999
dv = DataValidation(type="whole", operator="between", formula1="1", formula2="999", allow_blank=False)
dv.error = '请输入1到999之间的整数'
dv.errorTitle = '无效输入'
ws.add_data_validation(dv)
dv.add(ws["B1"])
# ---------------------------
# ② 基础模块计算区（蓝色区域）
# ---------------------------
current_row = 3
# 模块1：飞机采购成本（动态梯次模型）
ws.cell(row=current_row, column=1, value="一、飞机采购成本（动态梯次模型）")
current_row += 1
ws.cell(row=current_row, column=1, value="项目")
ws.cell(row=current_row, column=2, value="公式/数值")
current_row += 1
# 采购单价（根据飞机数量选择梯次单价：1-9架：239，10-19架：229，≥20架：219）
ws.cell(row=current_row, column=1, value="采购单价（万元）")
ws.cell(row=current_row, column=2, 
        value='=IF($B$1<10,239,IF($B$1<20,229,219))')
current_row += 1
# 采购总成本 = 飞机数量*采购单价
ws.cell(row=current_row, column=1, value="采购总成本（万元）")
ws.cell(row=current_row, column=2, 
        value='=$B$1*IF($B$1<10,239,IF($B$1<20,229,219))')
current_row += 2
# 模块2：机场基础设施成本（边际成本池）
ws.cell(row=current_row, column=1, value="二、机场基础设施成本（边际成本池）")
current_row += 1
ws.cell(row=current_row, column=1, value="设施类型")
ws.cell(row=current_row, column=2, value="触发条件")
ws.cell(row=current_row, column=3, value="单位成本（万元）")
ws.cell(row=current_row, column=4, value="总成本")
current_row += 1
# 基础机场包（≥1架）
ws.cell(row=current_row, column=1, value="基础机场包")
ws.cell(row=current_row, column=2, value="≥1架")
ws.cell(row=current_row, column=3, value=500)
ws.cell(row=current_row, column=4, value='=IF($B$1>=1,500,0)')
current_row += 1
# 扩容包A（≥5架）
ws.cell(row=current_row, column=1, value="扩容包A")
ws.cell(row=current_row, column=2, value="≥5架")
ws.cell(row=current_row, column=3, value=200)
ws.cell(row=current_row, column=4, value='=IF($B$1>=5,200,0)')
current_row += 1
# 扩容包B（≥10架）
ws.cell(row=current_row, column=1, value="扩容包B")
ws.cell(row=current_row, column=2, value="≥10架")
ws.cell(row=current_row, column=3, value=150)
ws.cell(row=current_row, column=4, value='=IF($B$1>=10,150,0)')
current_row += 1
# 智能调度系统（≥15架）
ws.cell(row=current_row, column=1, value="智能调度系统")
ws.cell(row=current_row, column=2, value="≥15架")
ws.cell(row=current_row, column=3, value=300)
ws.cell(row=current_row, column=4, value='=IF($B$1>=15,300,0)')
current_row += 1
# 基础设施总成本
ws.cell(row=current_row, column=1, value="基础设施总成本")
ws.cell(row=current_row, column=4, value="=SUM(D"+str(current_row-4)+":D"+str(current_row-1)+")")
current_row += 2
# 模块3：候机楼模块化成本
ws.cell(row=current_row, column=1, value="三、候机楼模块化成本")
current_row += 1
ws.cell(row=current_row, column=1, value="功能区")
ws.cell(row=current_row, column=2, value="面积系数")
ws.cell(row=current_row, column=3, value="基础成本（万元）")
ws.cell(row=current_row, column=4, value="边际成本率")
ws.cell(row=current_row, column=5, value="总成本")
current_row += 1
# 调控中心（固定）
ws.cell(row=current_row, column=1, value="调控中心")
ws.cell(row=current_row, column=2, value="固定")
ws.cell(row=current_row, column=3, value=80)
ws.cell(row=current_row, column=4, value="0%")
ws.cell(row=current_row, column=5, value=80)
current_row += 1
# 接待区
ws.cell(row=current_row, column=1, value="接待区")
ws.cell(row=current_row, column=2, value="1.2")
ws.cell(row=current_row, column=3, value=100)
ws.cell(row=current_row, column=4, value="5%/5架")
# =100 + INT(飞机数/5)*5%*100
ws.cell(row=current_row, column=5, value="=100+INT($B$1/5)*0.05*100")
current_row += 1
# 展示厅
ws.cell(row=current_row, column=1, value="展示厅")
ws.cell(row=current_row, column=2, value="1.5")
ws.cell(row=current_row, column=3, value=150)
ws.cell(row=current_row, column=4, value="3%/架")
ws.cell(row=current_row, column=5, value="=150*(1+0.03*$B$1)")
current_row += 1
# 活动中心
ws.cell(row=current_row, column=1, value="活动中心")
ws.cell(row=current_row, column=2, value="动态")
ws.cell(row=current_row, column=3, value=200)
ws.cell(row=current_row, column=4, value="10%/10架")
ws.cell(row=current_row, column=5, value="=200*CEILING($B$1/10,1)")
current_row += 1
# 候机楼总成本
ws.cell(row=current_row, column=1, value="候机楼总成本")
ws.cell(row=current_row, column=5, value="=SUM(E"+str(current_row-3)+":E"+str(current_row)+")")
current_row += 2
# 模块4：动态成本池（核心算法区）
ws.cell(row=current_row, column=1, value="四、动态成本池（核心算法区）")
current_row += 1
ws.cell(row=current_row, column=1, value="动态成本系数")
# LOOKUP公式：当飞机数≥5时下降10%，≥10时0.85，≥15时0.8
ws.cell(row=current_row, column=2, 
        value='=LOOKUP($B$1, {0,5,10,15}, {1,0.9,0.85,0.8})')
current_row += 1
ws.cell(row=current_row, column=1, value="最终总成本（万元）")
# 最终总成本 = (采购总成本 + 基础设施总成本 + 候机楼总成本)*动态成本系数
# 假设采购总成本在C采购模块第1模块里，写在此处引用对应单元格
# 采购总成本公式写在本模板中第模块1中，我们在此用单元格引用（假设采购总成本在B5+?）
# 这里我们依次引用：采购总成本在 cell B4（采购单价在B3，采购总成本在B4），
# 基础设施总成本在 D?（这里用 D17，候机楼总成本在 E?（用 E23）），动态系数在 B?（B25）
ws.cell(row=current_row, column=2, 
        value="=($B$1*IF($B$1<10,239,IF($B$1<20,229,219)) + D"+str(current_row-11)+"+E"+str(current_row-6)+")*B"+str(current_row-2))
current_row += 2
# ---------------------------
# ③ 增强模块
# ---------------------------
# A. 绿色【规模对照】
ws.cell(row=current_row, column=1, value="【规模对照】")
current_row += 1
ws.cell(row=current_row, column=1, value="对比规模")
ws.cell(row=current_row, column=2, value="采购成本")
ws.cell(row=current_row, column=3, value="基础设施")
ws.cell(row=current_row, column=4, value="候机楼成本")
ws.cell(row=current_row, column=5, value="总成本")
ws.cell(row=current_row, column=6, value="成本区间提示")
current_row += 1
# 对比规模：分别用5、10、15架计算
for scale, row in zip([5, 10, 15], range(current_row, current_row+3)):
    ws.cell(row=row, column=1, value=f"{scale}架")
    # 采购成本： = scale * IF(scale<10,239,IF(scale<20,229,219))
    ws.cell(row=row, column=2, 
            value=f"={scale}*IF({scale}<10,239,IF({scale}<20,229,219))")
    # 基础设施： =IF(scale>=1,500,0)+IF(scale>=5,200,0)+IF(scale>=10,150,0)+IF(scale>=15,300,0)
    ws.cell(row=row, column=3, 
            value=f"=IF({scale}>=1,500,0)+IF({scale}>=5,200,0)+IF({scale}>=10,150,0)+IF({scale}>=15,300,0)")
    # 候机楼成本： 同上模块3公式
    # 调控中心：80；接待区：=100+INT(scale/5)*0.05*100；展示厅：=150*(1+0.03*scale)；活动中心：=200*CEILING(scale/10,1)
    ws.cell(row=row, column=4, 
            value=f"=80+(100+INT({scale}/5)*0.05*100)+150*(1+0.03*{scale})+200*CEILING({scale}/10,1)")
    # 总成本：先合计，再乘以动态系数：动态系数=LOOKUP(scale,{0,5,10,15},{1,0.9,0.85,0.8})
    ws.cell(row=row, column=5, 
            value=f"=({scale}*IF({scale}<10,239,IF({scale}<20,229,219)) + (IF({scale}>=1,500,0)+IF({scale}>=5,200,0)+IF({scale}>=10,150,0)+IF({scale}>=15,300,0)) + (80+(100+INT({scale}/5)*0.05*100)+150*(1+0.03*{scale})+200*CEILING({scale}/10,1)))*LOOKUP({scale},{{0,5,10,15}},{{1,0.9,0.85,0.8}})")
    # 成本区间提示（示例逻辑）：若成本低于某个基准（此处用3000做示范）则“不经济”，否则“优化区间”
    ws.cell(row=row, column=6, 
            value="=IF(E"+str(row)+"<3000,\"规模不经济\",\"优化区间\")")
current_row += 4
# B. 红色【敏感度分析矩阵】
ws.cell(row=current_row, column=1, value="【敏感度分析矩阵】")
current_row += 1
# 标题行：规模\单价波动
ws.cell(row=current_row, column=1, value="规模\\单价波动")
ws.cell(row=current_row, column=2, value="+15%单价")
ws.cell(row=current_row, column=3, value="+10%单价")
ws.cell(row=current_row, column=4, value="基准值")
ws.cell(row=current_row, column=5, value="-10%单价")
ws.cell(row=current_row, column=6, value="-20%单价")
current_row += 1
# 这里以 5、10、15 架为例，基于上面【规模对照】的总成本，做比例调整
for scale, row in zip([5, 10, 15], range(current_row, current_row+3)):
    # 基准总成本公式，与【规模对照】中的公式类似
    base_formula = f"({scale}*IF({scale}<10,239,IF({scale}<20,229,219)) + (IF({scale}>=1,500,0)+IF({scale}>=5,200,0)+IF({scale}>=10,150,0)+IF({scale}>=15,300,0)) + (80+(100+INT({scale}/5)*0.05*100)+150*(1+0.03*{scale})+200*CEILING({scale}/10,1)))*LOOKUP({scale},{{0,5,10,15}},{{1,0.9,0.85,0.8}})"
    ws.cell(row=row, column=1, value=f"{scale}架成本")
    ws.cell(row=row, column=2, value=f"=({base_formula})*(1+15%)")
    ws.cell(row=row, column=3, value=f"=({base_formula})*(1+10%)")
    ws.cell(row=row, column=4, value=f"=({base_formula})")
    ws.cell(row=row, column=5, value=f"=({base_formula})*(1-10%)")
    ws.cell(row=row, column=6, value=f"=({base_formula})*(1-20%)")
current_row += 2
# C. 边际成本拐点提示（示例提示逻辑）
ws.cell(row=current_row, column=1, value="边际成本拐点提示：")
ws.cell(row=current_row, column=2, 
        value='=IF($B$1=5,"当前规模已达5架，建议增至10架享受成本跃迁",IF($B$1=10,"当前规模已达10架，建议增至15架享受成本跃迁",""))')
# ---------------------------
# ④ 条件格式示例（例如对成本区间自动变色，可根据需要调整）
# ---------------------------
# 此处示例对【规模对照】区域E列（总成本）设置条件格式：若总成本低于3000，则填充红色
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
rule = FormulaRule(formula=["E"+str(current_row-6)+"<3000"], fill=red_fill)
# 为简便只对【规模对照】中第一行设置（实际可根据区域循环设置）
ws.conditional_formatting.add("E"+str(current_row-6), rule)
# ---------------------------
# 保存文件
# ---------------------------
wb.save("增强版成本模板V1.xlsx")
print("Excel 文件 '增强版成本模板V1.xlsx' 已生成。")
