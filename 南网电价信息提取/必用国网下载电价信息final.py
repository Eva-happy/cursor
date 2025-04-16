import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import pdfplumber
import warnings
import re
import math
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import urllib3
import openpyxl
import io
import fitz  # PyMuPDF
import PIL.Image
import pyautogui
import subprocess
import win32com.client
import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
def setup_driver():
    """设置并返回WebDriver"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        driver_path = os.path.join(script_dir, "msedgedriver.exe")
        
        if not os.path.exists(driver_path):
            print(f"错误: msedgedriver.exe 不存在于 {driver_path}")
            return None
        
        edge_options = Options()
        edge_options.add_argument("--start-maximized")
        edge_options.add_argument('--disable-gpu')
        edge_options.add_argument('--no-sandbox')
        edge_options.add_argument('--disable-dev-shm-usage')
        edge_options.add_argument('--ignore-certificate-errors')
        # 添加以下参数来禁用日志
        edge_options.add_argument('--silent')
        edge_options.add_argument('--log-level=3')  # 只显示致命错误
        edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])  # 禁用 USB 设备检查日志
        
        service = Service(driver_path)
        driver = webdriver.Edge(service=service, options=edge_options)
        return driver
    except Exception as e:
        print(f"设置驱动程序失败: {str(e)}")
        return None
def wait_and_click_element(driver, element, use_js=False):
    """等待并点击元素"""
    try:
        element_text = "选择地区" if element.text.strip() == "上海" else element.text.strip()
        print(f"\n正在点击: {element_text}")
        
        # 尝试多种点击方法
        try_methods = [
            # 方法1：常规点击
            lambda: element.click(),
            # 方法2：JavaScript点击
            lambda: driver.execute_script("arguments[0].click();", element),
            # 方法3：先滚动再点击
            lambda: (
                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element),
                time.sleep(0.5),
                element.click()
            ),
            # 方法4：移除可能遮挡的元素后点击
            lambda: (
                driver.execute_script("""
                    var elements = document.getElementsByClassName('tab-content');
                    for(var i=0; i<elements.length; i++){
                        elements[i].style.pointerEvents = 'none';
                    }
                """),
                element.click()
            ),
            # 方法5：使用ActionChains
            lambda: ActionChains(driver).move_to_element(element).click().perform()
        ]
        
        # 依次尝试不同的点击方法
        for i, method in enumerate(try_methods, 1):
            try:
                print(f"尝试第{i}种点击方法...")
                method()
                print(f"成功点击: {element_text}")
                return True
            except Exception as e:
                print(f"第{i}种方法失败: {str(e)}")
                continue
        
        print("所有点击方法都失败")
        return False
        
    except Exception as e:
        print(f"点击元素失败: {str(e)}")
        return False
def find_element_with_retry(driver, selectors, timeout=5):
    """使用多个选择器尝试查找元素"""
    for selector in selectors:
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, selector))
            )
            if element.is_displayed():
                return element
        except:
            continue
    return None
def get_visible_elements(driver, selector, timeout=3):  # 减少默认等待时间
    """获取所有可见的元素"""
    try:
        elements = WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((By.XPATH, selector))
        )
        return [e for e in elements if e.is_displayed()]
    except:
        return []
def extract_element_text(element):
    """提取元素文本"""
    try:
        text = element.text.strip()
        if element.tag_name == 'tr':
            cells = element.find_elements(By.CSS_SELECTOR, 'div.cell')
            cell_texts = [cell.text.strip() for cell in cells if cell.text.strip()]
            if cell_texts:
                text = ' | '.join(cell_texts)
        return text
    except:
        return ""

def get_clickable_element(element):
    """获取可点击的元素"""
    try:
        return element.find_element(By.TAG_NAME, "a")
    except:
        return element
    
def get_provinces(driver):
    """获取省份列表"""
    try:
        # 点击地区选择器
        region_selectors = [
            "//div[@id='city_select']//a[contains(@class, 'current')]",
            "//div[contains(@class, 'region')]//a[contains(@class, 'current')]",
            "//a[contains(@class, 'current fsize16')]",
            "//div[@data-v-07831be2]//a[contains(@class, 'current')]"
        ]
        
        region_element = find_element_with_retry(driver, region_selectors)
        if not region_element or not wait_and_click_element(driver, region_element):
            print("点击地区选择器失败")
            return []
        
        time.sleep(2)
        
        # 获取省份列表
        province_selectors = [
            "//a[contains(@class, 'f66 fsize14')]",
            "//div[contains(@class, 'province-list')]//a",
            "//div[contains(@class, 'region-list')]//a"
        ]
        
        provinces = []
        for selector in province_selectors:
            elements = get_visible_elements(driver, selector)
            for element in elements:
                text = extract_element_text(element)
                if text and text != '省份' and text not in [p['text'] for p in provinces]:
                    provinces.append({"text": text, "element": element})
        
        return provinces
    except Exception as e:
        print(f"获取省份列表失败: {str(e)}")
        return []
def get_cities(driver):
    """获取城市列表"""
    print("\n正在等待城市列表加载...")
    time.sleep(2)  # 增加等待时间
    
    # 等待对话框出现
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div[data-v-40fe627e].tab-con-box'))
        )
        print("城市选择框已加载")
    except Exception as e:
        print(f"等待城市选择框超时: {str(e)}")
    
    # 更多的选择器
    city_selectors = [
        # 基本选择器
        "//div[contains(@class, 'tab-con-box')]//li[contains(@class, 'tab-con-box-li')]",
        "//ul[contains(@class, 'tab-con-box-ul')]//li[contains(@class, 'tab-con-box-li')]",
        "//div[contains(@class, 'el-dialog_body')]//li[contains(@class, 'tab-con-box-li')]",
        # 新增选择器
        "//div[contains(@class, 'el-dialog_body')]//li",
        "//div[contains(@class, 'city-list')]//li",
        "//div[contains(@class, 'tab-con-box')]//li",
        "//div[contains(@class, 'dialog')]//li[contains(@class, 'tab-con-box-li')]",
        "//div[contains(@class, 'el-dialog')]//ul//li"
    ]
    
    cities = []
    for selector in city_selectors:
        try:
            print(f"\n尝试使用选择器: {selector}")
            elements = WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located((By.XPATH, selector))
            )
            visible_elements = [e for e in elements if e.is_displayed()]
            
            if visible_elements:
                print(f"找到 {len(visible_elements)} 个可见元素")
                for element in visible_elements:
                    text = extract_element_text(element)
                    if text and text != '请选择' and text not in [c['text'] for c in cities]:
                        cities.append({"text": text, "element": element})
                if cities:
                    print(f"成功获取到 {len(cities)} 个城市")
                    break
        except Exception as e:
            print(f"使用选择器 {selector} 失败: {str(e)}")
            continue
    
    if not cities:
        print("\n尝试使用JavaScript获取城市列表...")
        try:
            cities_js = driver.execute_script("""
                var cities = [];
                var elements = document.querySelectorAll('div[data-v-40fe627e].tab-con-box ul[data-v-40fe627e].tab-con-box-ul li[data-v-40fe627e].tab-con-box-li');
                elements.forEach(function(el) {
                    if (el.textContent && el.textContent.trim() && el.offsetParent !== null) {
                        cities.push({
                            text: el.textContent.trim(),
                            element: el
                        });
                    }
                });
                return cities;
            """)
            
            if cities_js:
                for city in cities_js:
                    if city['text'] != '请选择' and city['text'] not in [c['text'] for c in cities]:
                        cities.append({"text": city['text'], "element": city['element']})
                print(f"通过JavaScript找到 {len(cities)} 个城市")
        except Exception as e:
            print(f"JavaScript方法失败: {str(e)}")
    
    if cities:
        print("\n可选城市列表:")
        for i, city in enumerate(cities, 1):
            print(f"{i}. {city['text']}")
    else:
        print("\n警告：未找到任何城市")
        # 打印当前页面结构以供调试
        try:
            container = driver.find_element(By.CSS_SELECTOR, 'div[data-v-40fe627e].tab-con-box')
            print("\n城市容器HTML:")
            print(container.get_attribute('innerHTML'))
        except:
            print("\n无法获取城市容器HTML")
            print("\n完整页面源代码:")
            print(driver.page_source)
    
    return cities
def get_districts(driver):
    """获取区县列表"""
    print("\n正在加载区县列表，请稍候...")
    time.sleep(2)  # 增加等待时间
    
    # 等待区县选择框出现
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div[data-v-40fe627e].tab-con-box'))
        )
        print("区县选择框已加载")
    except Exception as e:
        print(f"等待区县选择框超时: {str(e)}")
    
    # 区县选择器列表
    district_selectors = [
        # 基本选择器
        "//div[@data-v-40fe627e and @class='tab-con-box']//ul[@data-v-40fe627e and @class='tab-con-box-ul']//li[@data-v-40fe627e and @class='tab-con-box-li']",
        # 备用选择器
        "//div[contains(@class, 'tab-con-box')]//li[contains(@class, 'tab-con-box-li')]",
        "//ul[contains(@class, 'tab-con-box-ul')]//li[contains(@class, 'tab-con-box-li')]",
        "//div[contains(@class, 'tab-con-box')]//li",
        "//ul[contains(@class, 'tab-con-box-ul')]//li"
    ]
    
    districts = []
    for selector in district_selectors:
        try:
            print(f"\n尝试使用选择器: {selector}")
            elements = WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located((By.XPATH, selector))
            )
            visible_elements = [e for e in elements if e.is_displayed()]
            
            if visible_elements:
                print(f"找到 {len(visible_elements)} 个可见元素")
                for element in visible_elements:
                    text = extract_element_text(element)
                    if text and text != '请选择' and text not in [d['text'] for d in districts]:
                        districts.append({"text": text, "element": element})
                if districts:
                    print(f"成功获取到 {len(districts)} 个区县")
                    break
        except Exception as e:
            print(f"使用选择器 {selector} 失败: {str(e)}")
            continue
    
    if not districts:
        print("\n尝试使用JavaScript获取区县列表...")
        try:
            districts_js = driver.execute_script("""
                var districts = [];
                var elements = document.querySelectorAll('div[data-v-40fe627e].tab-con-box ul[data-v-40fe627e].tab-con-box-ul li[data-v-40fe627e].tab-con-box-li');
                elements.forEach(function(el) {
                    if (el.textContent && el.textContent.trim() && el.offsetParent !== null) {
                        districts.push({
                            text: el.textContent.trim(),
                            element: el
                        });
                    }
                });
                return districts;
            """)
            
            if districts_js:
                for district in districts_js:
                    if district['text'] != '请选择' and district['text'] not in [d['text'] for d in districts]:
                        districts.append({"text": district['text'], "element": district['element']})
                print(f"通过JavaScript找到 {len(districts)} 个区县")
        except Exception as e:
            print(f"JavaScript方法失败: {str(e)}")
    
    if districts:
        print("\n可选区县列表:")
        for i, district in enumerate(districts, 1):
            print(f"{i}. {district['text']}")
    else:
        print("\n警告：未找到任何区县")
        # 打印当前页面结构以供调试
        try:
            container = driver.find_element(By.CSS_SELECTOR, 'div[data-v-40fe627e].tab-con-box')
            print("\n区县容器HTML:")
            print(container.get_attribute('innerHTML'))
        except:
            print("\n无法获取区县容器HTML")
            print("\n完整页面源代码:")
            print(driver.page_source)
    
    return districts
def get_projects(driver):
    """获取项目列表"""
    try:
        print("\n正在获取电价标准项目列表...")
        
        # 等待表格加载
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.el-table__body-wrapper'))
        )
        time.sleep(1)  # 等待表格内容完全加载
        
        # 获取所有表格行
        rows = driver.find_elements(By.CSS_SELECTOR, '.el-table__body-wrapper .el-table__row')
        
        projects = []
        seen_texts = set()
        
        for row in rows:
            try:
                # 获取行中的所有单元格
                cells = row.find_elements(By.CSS_SELECTOR, '.cell')
                
                # 获取第一个单元格的文本（通常是项目名称）
                if cells:
                    project_text = cells[0].text.strip()
                    if project_text and project_text not in seen_texts:
                        seen_texts.add(project_text)
                        # 使用整个行作为元素，因为展开箭头在行级别
                        projects.append({"text": project_text, "element": row})
            except:
                continue
        
        if projects:
            print(f"\n成功获取到 {len(projects)} 个项目")
            return projects
        
        print("\n未找到任何项目，尝试刷新页面...")
        driver.refresh()
        time.sleep(2)
        
        # 再次尝试获取项目
        rows = driver.find_elements(By.CSS_SELECTOR, '.el-table__body-wrapper .el-table__row')
        for row in rows:
            try:
                cells = row.find_elements(By.CSS_SELECTOR, '.cell')
                if cells:
                    project_text = cells[0].text.strip()
                    if project_text and project_text not in seen_texts:
                        seen_texts.add(project_text)
                        projects.append({"text": project_text, "element": row})
            except:
                continue
        
        if projects:
            print(f"\n刷新后成功获取到 {len(projects)} 个项目")
        else:
            print("\n刷新后仍未找到任何项目")
        
        return projects
        
    except Exception as e:
        print(f"\n获取项目列表失败: {str(e)}")
        return []

def get_subprojects(driver, project_text):
    """获取子项目列表"""
    try:
        # 等待展开行加载
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "tr.el-table__expanded-row"))
        )
        time.sleep(1)  # 等待内容加载
        
        # 直接获取所有链接
        links = driver.find_elements(By.XPATH, "//tr[contains(@class, 'el-table__expanded-row')]//div[contains(@class, 'cell')]//a")
        
        if not links:
            print("\n展开行中没有找到链接")
            return []
        
        subprojects = []
        seen_texts = set()  # 用于去重
        
        for link in links:
            text = link.text.strip()
            if (text and 
                text not in seen_texts and 
                not text.startswith("序号") and 
                not text.startswith("文号") and
                not text.startswith("发布日期") and
                not text.startswith("实施日期") and
                "展开" not in text and
                len(text) > 1):
                seen_texts.add(text)
                subprojects.append({"text": text, "element": link})
        
        if subprojects:
            print(f"\n成功获取到 {len(subprojects)} 个子项目")
            # 打印所有子项目名称
            print("\n子项目列表:")
            for i, project in enumerate(subprojects, 1):
                print(f"{i}. {project['text']}")
        else:
            print("\n警告：未找到有效的子项目")
            
        return subprojects
        
    except Exception as e:
        print(f"\n获取子项目列表失败: {str(e)}")
        return []

def display_menu(items, title="请选择："):
    """显示菜单并获取用户选择"""
    while True:
        print(f"\n{title}")
        for i, item in enumerate(items, 1):
            print(f"{i}. {item['text']}")
        
        try:
            choice = int(input("\n请输入编号: "))
            if 1 <= choice <= len(items):
                selected_item = items[choice-1]
                print("\n" + "="*50)
                print(f"您选择了: {selected_item['text']}")
                print("="*50 + "\n")
                return selected_item
            print("无效的选择，请重试")
        except ValueError:
            print("请输入有效的数字")
    
def check_and_process_content(driver, project_name=None, is_yunnan=False):
    """检查并处理页面内容"""
    try:
        if is_yunnan:
            # 云南省特殊处理
            print("等待页面完全加载...")
            time.sleep(3)  # 给页面更多加载时间
            
            # 直接使用PDF下载链接
            pdf_url = "https://www.95598.cn/omg-static//omg-static/99302251957052709409300507214865.pdf"
            print("使用PDF下载链接:", pdf_url)
            
            # 使用下载方法
            pdf_file = download_pdf(pdf_url, project_name)
            if pdf_file:
                print("PDF下载成功，开始处理内容...")
                all_pages_data = extract_table_from_pdf(pdf_file, is_yunnan=True)
                if all_pages_data:
                    print("成功提取PDF数据，开始生成Excel文件...")
                    excel_file = write_to_excel(all_pages_data, project_name, is_yunnan=True)
                    if excel_file:
                        print("\n开始处理图片复制粘贴...")
                        # 直接从网页复制图片到Excel
                        copy_image_from_pdf_to_excel(driver, excel_file, [1, 2, 3])
                    return excel_file
                else:
                    print("PDF数据提取失败")
                    return None
            else:
                print("PDF下载失败")
                return None
        else:
            # 其他省份的处理逻辑
            try:
                wait = WebDriverWait(driver, 2)

                # 首先尝试查找普通的PDF iframe
                try:
                    iframe = wait.until(
                        EC.presence_of_element_located((By.XPATH, "//iframe[contains(@src, 'downloadFiles')]"))
                    )
                    pdf_url = iframe.get_attribute('src')
                except TimeoutException:
                # 如果找不到普通的PDF iframe，尝试查找并处理带链接的iframe
                    print("等待iframe加载...")
                    iframe = wait.until(
                        EC.presence_of_element_located((By.TAG_NAME, "iframe"))
                    )
                    driver.switch_to.frame(iframe)
                
                    print("查找并点击PDF链接...")
                    link = wait.until(
                        EC.element_to_be_clickable((By.TAG_NAME, "a"))
                    )
                    pdf_url = link.get_attribute('href')
                    
                    # 使用JavaScript点击链接
                    driver.execute_script("arguments[0].click();", link)
                    
                    # 切回主文档
                    driver.switch_to.default_content()
                    
                    # 等待新页面加载
                    print("等待PDF页面加载...")
                    time.sleep(2)  # 给页面一些加载时间
                    
                    # 获取当前窗口句柄
                    current_window = driver.current_window_handle
                    
                    # 切换到新窗口（如果有）
                    all_windows = driver.window_handles
                    for window in all_windows:
                        if window != current_window:
                            driver.switch_to.window(window)
                            break
            
                # 处理PDF内容
                all_pages_data = process_pdf_content(driver, pdf_url, project_name)
                if all_pages_data:
                    # 写入Excel
                    excel_file = write_to_excel(all_pages_data, project_name, is_yunnan=False)
                    return excel_file
                return None
            
            except TimeoutException:
                # 处理在线预览内容
                page_data = process_preview_content(driver)
                if page_data:
                    # 将单页数据包装成列表
                    all_pages_data = [page_data]
                    # 写入Excel
                    excel_file = write_to_excel(all_pages_data, project_name, is_yunnan=False)
                    return excel_file
                return None
            
    except Exception as e:
        print(f"处理内容失败: {str(e)}")
        import traceback
        traceback.print_exc()  # 打印完整的错误堆栈
        return None

    """处理在线预览内容"""
    try:
        print("\n开始处理在线预览内容...")
        
        # 使用显式等待等待页面加载完成
        wait = WebDriverWait(driver, 2)
        
        # 等待标题元素加载
        print("等待页面加载...")
        title = ""
        try:
            # 尝试多种可能的标题选择器
            title_selectors = [
                "//h1",
                "//div[contains(@class, 'title')]",
                "//div[contains(@class, 'header')]/div",
                "//div[contains(text(), '深圳供电局有限公司')]"
            ]
            
            for selector in title_selectors:
                try:
                    title_element = wait.until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )
                    title = title_element.text.strip()
                    if title:
                        print(f"找到标题: {title}")
                        break
                except TimeoutException:
                    continue
                except Exception:
                    continue
            
            if not title:
                print("未找到标题，使用默认值")
                title = "深圳供电局有限公司电价信息"
        except Exception as e:
            print(f"获取标题失败: {str(e)}")
            title = "深圳供电局有限公司电价信息"
        
        # 获取执行时间
        print("获取执行时间...")
        execution_time = ""
        try:
            time_selectors = [
                "//*[contains(text(), '执行时间')]",
                "//div[contains(text(), '年') and contains(text(), '月')]",
                "//span[contains(text(), '执行时间')]"
            ]
            
            for selector in time_selectors:
                try:
                    time_element = wait.until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )
                    execution_time = time_element.text.strip()
                    if execution_time:
                        print(f"找到执行时间: {execution_time}")
                        break
                except TimeoutException:
                    continue
                except Exception:
                    continue
        except Exception as e:
            print(f"获取执行时间失败: {str(e)}")
        
        # 获取表格数据
        print("获取表格数据...")
        table_data = []
        try:
            # 等待表格加载
            table = wait.until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            
            # 等待表格内容加载完成
            wait.until(
                lambda d: len(d.find_elements(By.TAG_NAME, "tr")) > 0
            )
            
            # 获取所有表格
            tables = driver.find_elements(By.TAG_NAME, "table")
            print(f"找到 {len(tables)} 个表格")
            
            if tables:
                # 使用第一个表格
                main_table = tables[0]
                
                # 获取所有行（包括表头和数据行）
                rows = main_table.find_elements(By.TAG_NAME, "tr")
                print(f"找到 {len(rows)} 行数据")
                
                for row in rows:
                    # 尝试获取表头单元格
                    header_cells = row.find_elements(By.TAG_NAME, "th")
                    if header_cells:
                        row_data = [cell.text.strip() for cell in header_cells]
                    else:
                        # 获取数据单元格
                        cells = row.find_elements(By.TAG_NAME, "td")
                        row_data = [cell.text.strip() for cell in cells]
                    
                    if any(row_data):  # 只添加非空行
                        table_data.append(row_data)
                        print(f"添加行: {row_data}")
            
            if not table_data:
                print("未找到表格数据，尝试其他方法...")
                try:
                    # 等待表格主体加载
                    table_body = wait.until(
                        EC.presence_of_element_located((By.CLASS_NAME, "el-table__body-wrapper"))
                    )
                    
                    # 等待表头加载
                    table_header = wait.until(
                        EC.presence_of_element_located((By.CLASS_NAME, "el-table__header-wrapper"))
                    )
                    
                    # 获取表头
                    header_rows = table_header.find_elements(By.TAG_NAME, "tr")
                    for row in header_rows:
                        cells = row.find_elements(By.TAG_NAME, "th")
                        row_data = [cell.text.strip() for cell in cells]
                        if any(row_data):
                            table_data.append(row_data)
                            print(f"添加表头行: {row_data}")
                    
                    # 获取数据行
                    rows = table_body.find_elements(By.CLASS_NAME, "el-table__row")
                    for row in rows:
                        cells = row.find_elements(By.CLASS_NAME, "el-table__cell")
                        row_data = [cell.text.strip() for cell in cells]
                        if any(row_data):
                            table_data.append(row_data)
                            print(f"添加数据行: {row_data}")
                except TimeoutException:
                    print("等待表格元素超时")
                except Exception as e:
                    print(f"获取表格数据失败: {str(e)}")
        
        except TimeoutException:
            print("等待表格加载超时")
        except Exception as e:
            print(f"获取表格数据失败: {str(e)}")
        
        if not table_data:
            print("未能获取到任何表格数据")
            return None
        
        # 获取注释内容
        print("获取注释内容...")
        notes = []
        try:
            # 尝试多种可能的注释选择器
            note_selectors = [
                "//*[contains(text(), '注：')]",
                "//*[contains(text(), '注:')]",
                "//div[starts-with(text(), '注：')]",
                "//div[starts-with(text(), '注:')]",
                "//p[contains(text(), '注：')]",
                "//*[contains(text(), '备注：')]",
                "//*[contains(text(), '备注:')]"
            ]
            
            for selector in note_selectors:
                try:
                    note_elements = driver.find_elements(By.XPATH, selector)
                    for element in note_elements:
                        note_text = element.text.strip()
                        if note_text and note_text not in notes:
                            notes.append(note_text)
                            print(f"找到注释: {note_text}")
                except Exception:
                    continue
        except Exception as e:
            print(f"获取注释失败: {str(e)}")
        
        # 创建page_data字典
        print("创建数据字典...")
        page_data = {
            'title': title,
            'subtitle': execution_time,
            'table': table_data,
            'notes': notes
        }
        
        return page_data
        
    except Exception as e:
        print(f"处理在线预览内容失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

    """处理第一页数据（主电价表）"""
    try:
        if not page_data or 'table' not in page_data:
            print("没有找到表格数据")
            return None, None, None
            
        table_data = page_data['table']
        if not table_data:
            print("表格数据为空")
            return None, None, None
            
        # 提取注释
        notes = page_data.get('notes', [])
        
        # 直接将原始表格数据转换为DataFrame，不做任何处理
        df = pd.DataFrame(table_data)
        
        return df, notes, None
    except Exception as e:
        print(f"处理第一页数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None, None

    """处理第二页数据（使用说明）"""
    try:
        if not page_data or 'table' not in page_data:
            print("没有找到表格数据")
            return None
            
        table_data = page_data['table']
        if not table_data:
            print("表格数据为空")
            return None
            
        # 直接将原始表格数据转换为DataFrame，不做任何处理
        df = pd.DataFrame(table_data)
        
        return df
    except Exception as e:
        print(f"处理第二页数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

    """处理PDF内容
    Args:
        driver: WebDriver对象
        pdf_url: PDF文件的URL
        project_name: 项目名称，用于文件命名
    """
    try:
        if not pdf_url:
            return None
            
        # 下载PDF文件
        pdf_file = download_pdf(pdf_url, project_name)
        if not pdf_file:
            return None
                
        # 检查PDF文件是否存在且大小正常
        if not os.path.exists(pdf_file) or os.path.getsize(pdf_file) == 0:
            return None
            
        # 提取表格数据
        all_pages_data = extract_table_from_pdf(pdf_file)
        if not all_pages_data:
            return None
            
        # 直接返回提取的数据，而不是Excel文件路径
        return all_pages_data
            
    except Exception as e:
        print(f"处理PDF内容失败: {str(e)}")
        return None
def download_pdf(pdf_url, project_name=None):
    """下载PDF文件
    Args:
        pdf_url: PDF文件的URL
        project_name: 项目名称，用于文件命名
    """
    try:
        # 创建下载目录
        if not os.path.exists('downloads'):
            os.makedirs('downloads')
        
        # 生成基础文件名
        base_filename = project_name if project_name else f'电价表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        
        # 检查是否存在同名文件，并添加版本号
        version = 1
        while True:
            pdf_file = os.path.join('downloads', f'{base_filename}V{version}.pdf')
            if not os.path.exists(pdf_file):
                break
            version += 1
        
        # 下载PDF
        print(f"正在下载PDF到: {pdf_file}")
        
        # 设置请求头
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'application/pdf,*/*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Referer': 'https://95598.cn/'
        }
        
        # 禁用SSL验证警告
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # 使用session来处理请求
        session = requests.Session()
        session.verify = False  # 禁用SSL验证
        
        # 设置超时时间
        response = session.get(pdf_url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            # 检查内容类型和文件大小
            content_type = response.headers.get('content-type', '').lower()
            content_length = len(response.content)
            
            if content_length == 0:
                print("下载的内容为空")
                return None
            
            # 保存文件
            with open(pdf_file, 'wb') as f:
                f.write(response.content)
            
            # 验证保存的文件
            if os.path.exists(pdf_file) and os.path.getsize(pdf_file) > 0:
                if 'application/pdf' in content_type or response.content.startswith(b'%PDF'):
                    print("PDF下载完成")
                else:
                    print(f"警告：文件已保存，但内容类型不是PDF (Content-Type: {content_type})")
                return pdf_file
            else:
                print("保存的文件为空或不存在")
                return None
        else:
            print(f"下载失败，状态码: {response.status_code}")
            return None
    except requests.Timeout:
        print("下载超时")
        return None
    except requests.RequestException as e:
        print(f"下载请求失败: {str(e)}")
        return None
    except Exception as e:
        print(f"下载PDF时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

    """使用PyMuPDF从本地PDF提取图片,专门用于云南省"""
    try:
        doc = fitz.open(pdf_path)
        if 0 <= page_num - 1 < doc.page_count:
            page = doc[page_num - 1]
            image_list = page.get_images()
            
            if image_list:
                print(f"第 {page_num} 页发现 {len(image_list)} 张图片")
                result = []
                
                for img_index, img in enumerate(image_list, start=1):
                    try:
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        
                        if base_image:
                            # 直接使用原始图片数据
                            image_bytes = base_image["image"]
                            image_ext = base_image["ext"]
                            
                            # 创建一个临时的内存文件
                            img_stream = io.BytesIO()
                            # 写入图片数据
                            img_stream.write(image_bytes)
                            # 将指针移到开始位置
                            img_stream.seek(0)
                            
                            result.append({
                                'stream': img_stream,
                                'width': base_image.get("width", 800),
                                'height': base_image.get("height", 600),
                                'ext': image_ext
                            })
                            print(f"成功提取图片 {img_index}，格式：{image_ext}")
                    except Exception as e:
                        print(f"处理图片 {img_index} 时出错: {str(e)}")
                        continue
                
                return result
            else:
                print(f"第 {page_num} 页没有找到图片")
                return []
    except Exception as e:
        print(f"提取图片时出错: {str(e)}")
        return []
    finally:
        if 'doc' in locals():
            doc.close()

    """使用PyMuPDF从本地PDF提取文本和图片，专门用于云南省"""
    try:
        print(f"\n开始处理PDF文件: {pdf_path}")
        all_pages_data = []
        doc = fitz.open(pdf_path)
        total_pages = doc.page_count
        print(f"PDF共有 {total_pages} 页")
            
        for page_num in range(total_pages):
            try:
                print(f"\n处理第 {page_num + 1} 页...")
                page = doc[page_num]
                page_data = {}
                    
                # 提取图片（前三页）
                if page_num < 3:
                    print(f"正在提取第 {page_num + 1} 页的图片...")
                    image_data = extract_images_from_pdf_yunnan(pdf_path, page_num + 1)
                    if image_data:
                        print(f"成功提取到 {len(image_data)} 张图片")
                        page_data['images'] = image_data
                    else:
                        print("未找到图片")
                
                # 提取文本
                text = page.get_text()
                print(f"提取到的文本长度: {len(text)}")
                print(f"文本前100个字符预览: {text[:100]}")
                
                if not text:
                    print(f"第 {page_num + 1} 页文本为空")
                    if 'images' in page_data:
                        all_pages_data.append(page_data)
                    continue
                
                # 分行处理文本
                lines = text.split('\n')
                lines = [line.strip() for line in lines if line.strip()]
                print(f"处理后的文本行数: {len(lines)}")
                
                if not lines:
                    print(f"第 {page_num + 1} 页没有文本行")
                    if 'images' in page_data:
                        all_pages_data.append(page_data)
                    continue
                
                # 根据页码不同处理文本内容
                if page_num < 3:  # 前三页：保留表格结构
                    print("使用前三页的处理逻辑")
                    # 提取标题（第一行）
                    page_data['title'] = lines[0]
                    print(f"标题: {page_data['title']}")
                        
                    # 提取执行时间（如果有）
                    for line in lines[1:3]:
                        if '执行时间' in line:
                            page_data['subtitle'] = line
                            print(f"执行时间: {line}")
                            break
                    
                    # 使用正则表达式清理文本
                    cleaned_lines = []
                    current_line = []
                    
                    # 只处理第一页和第二页的表格数据
                    if page_num < 2:
                        for line in lines[1:]:  # 跳过标题
                            line = line.strip()
                            if line:
                                current_line.append(line)
                                if not re.search(r'[a-zA-Z0-9]$', line):  # 如果行不以字母或数字结尾
                                    if current_line:
                                        cleaned_lines.append([' '.join(current_line)])
                                        current_line = []
                        
                        if current_line:  # 添加最后一行
                            cleaned_lines.append([' '.join(current_line)])
                    
                    # 设置表格数据
                    if cleaned_lines:
                        page_data['table'] = cleaned_lines
                        print(f"表格数据行数: {len(cleaned_lines)}")
                    else:
                        page_data['table'] = []
                        print("表格数据为空")
                
                else:  # 后三页：按段落分割
                    print("使用后三页的处理逻辑")
                    paragraphs = []
                    current_paragraph = []
                        
                    # 直接从第一行开始处理，不跳过标题
                    for i, line in enumerate(lines):
                        line = line.strip()
                        if line:
                            # 如果是新段落的开始（数字序号开头）
                            if re.match(r'^\d+[\.\、]', line):
                                if current_paragraph:
                                    # 合并当前段落并添加
                                    paragraphs.append([' '.join(current_paragraph)])
                                current_paragraph = [line]
                            else:
                                current_paragraph.append(line)
                                # 检查是否是段落结束
                                next_line = lines[i + 1] if i + 1 < len(lines) else None
                                if (line.endswith('。') and 
                                    (not next_line or re.match(r'^\d+[\.\、]', next_line))):
                                    if current_paragraph:
                                        paragraphs.append([' '.join(current_paragraph)])
                                        current_paragraph = []
                        
                    if current_paragraph:  # 添加最后一个段落
                        paragraphs.append([' '.join(current_paragraph)])
                        
                    if paragraphs:
                        page_data['table'] = paragraphs
                        print(f"段落数: {len(paragraphs)}")
                    
                # 提取注释
                notes = []
                if page_num < 3:  # 只在前三页提取注释
                    for line in lines:
                        if any(line.startswith(prefix) for prefix in ['注：', '注:', '备注：', '备注:', '注释说明：']):
                            notes.append(line)
                    
                    if notes:
                        page_data['notes'] = notes
                        print(f"注释数: {len(notes)}")
                    else:
                        page_data['notes'] = []
                        print("注释数: 0")
                else:
                    page_data['notes'] = []
                    print("注释数: 0")
                    
                # 提取单位信息
                for line in lines:
                    if '单位' in line:
                        page_data['unit'] = line
                        print(f"单位信息: {line}")
                        break
                    
                all_pages_data.append(page_data)
                print(f"第 {page_num + 1} 页处理完成")
                    
            except Exception as e:
                print(f"处理第 {page_num + 1} 页时出错: {str(e)}")
                continue
        
        print(f"\n总共处理了 {len(all_pages_data)} 页数据")
        return all_pages_data
    
    except Exception as e:
        print(f"提取文本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        if 'doc' in locals():
            doc.close()

    """从PDF中提取表格数据、文本内容和图片"""
    try:
        if is_yunnan:
            # 云南省特殊处理
            return extract_text_from_yunnan_pdf (pdf_path)
        else:
            # 其他省份的处理逻辑
            try:
                all_pages_data = []
                with pdfplumber.open(pdf_path) as pdf:
                    for page_num, page in enumerate(pdf.pages, 1):
                        try:
                            print(f"\n处理第 {page_num} 页...")
                            page_data = {}
                            
                            # 提取文本
                            text = page.extract_text()
                            if not text:
                                print(f"第 {page_num} 页文本为空")
                                continue
                            
                            lines = text.split('\n')
                            if not lines:
                                print(f"第 {page_num} 页没有文本行")
                                continue
                            
                            # 提取标题
                            title_line = lines[0].strip() if lines else None
                            if title_line:
                                page_data['title'] = title_line
                                print(f"标题: {title_line}")
                            
                            # 提取执行时间
                            for line in lines[1:3]:
                                if '执行时间' in line:
                                    page_data['subtitle'] = line.strip()
                                    print(f"执行时间: {line.strip()}")
                                    break
                            
                            # 提取表格
                            tables = page.extract_tables()
                            if tables:
                                main_table = tables[0]  # 使用第一个表格
                                if main_table and len(main_table) > 0:
                                    print(f"表格列数: {len(main_table[0])}")
                                    print(f"表头内容: {main_table[0]}")
                                    
                                    # 清理表格数据（移除空行和空列）
                                    cleaned_table = []
                                    for row in main_table:
                                        if any(cell.strip() if isinstance(cell, str) else cell for cell in row):
                                            cleaned_row = [
                                                cell.strip() if isinstance(cell, str) else cell
                                                for cell in row
                                            ]
                                            cleaned_table.append(cleaned_row)
                                    
                                    if cleaned_table:
                                        page_data['table'] = cleaned_table
                                    else:
                                        print("表格数据清理后为空")
                                else:
                                    print("表格结构无效")
                            else:
                                print("未找到表格")
                            
                            # 提取注释
                            notes = []
                            current_note = ""
                            in_notes = False
                            
                            for line in lines:
                                line = line.strip()
                                # 增加对"备注："的识别
                                if line.startswith('注：') or line.startswith('注:') or line.startswith('注释说明：') or line.startswith('备注：') or line.startswith('备注:'):
                                    if current_note:
                                        notes.append(current_note.strip())
                                    current_note = line if not (line.startswith('注释说明：') or line == '备注：' or line == '备注:') else ""
                                    in_notes = True
                                elif in_notes and line and not line.startswith('执行时间'):
                                    try:
                                        if re.match(r'^\d+[\.\、]', line):
                                            if current_note:
                                                notes.append(current_note.strip())
                                            current_note = line
                                        else:
                                            current_note = (current_note + " " + line) if current_note else line
                                    except Exception as e:
                                        print(f"处理注释行时出错: {str(e)}")
                                        continue
                            
                            if current_note:
                                notes.append(current_note.strip())
                            
                            # 处理注释内容
                            processed_notes = []
                            if notes:
                                # 移除第一个注释的前缀（"注："或"备注："）
                                if notes[0].startswith(('注：', '注:', '备注：', '备注:')):
                                    for prefix in ['注：', '注:', '备注：', '备注:']:
                                        if notes[0].startswith(prefix):
                                            notes[0] = notes[0][len(prefix):].strip()
                                            break
                                processed_notes = [note.strip() for note in notes if note.strip()]
                            
                            if processed_notes:
                                page_data['notes'] = processed_notes
                                print(f"找到 {len(processed_notes)} 条注释")
                            
                            # 提取单位信息
                            for line in lines:
                                if '单位' in line:
                                    page_data['unit'] = line.strip()
                                    print(f"单位信息: {line.strip()}")
                                    break
                            
                            # 只有当页面包含必要的数据时才添加
                            if 'table' in page_data:
                                all_pages_data.append(page_data)
                                print(f"第 {page_num} 页处理完成")
                            else:
                                print(f"第 {page_num} 页缺少必要的表格数据，已跳过")
                            
                        except Exception as e:
                            print(f"处理第 {page_num} 页时出错: {str(e)}")
                            continue
                
                if not all_pages_data:
                    print("未能从PDF中提取到有效数据")
                    return None
                
                print(f"共处理了 {len(all_pages_data)} 页有效数据")
                return all_pages_data
                
            except Exception as e:
                print(f"提取表格失败: {str(e)}")
                import traceback
                traceback.print_exc()
                return None
    
    except Exception as e:
        print(f"提取表格失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

    """合并空白单元格与相邻的有内容单元格"""
    def get_merged_ranges():
        try:
            return list(worksheet.merged_cells.ranges)
        except Exception:
            return []
    
    def is_cell_in_merged_range(row, col, ranges):
        try:
            return any(r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col 
                      for r in ranges)
        except Exception:
            return False
    
    def try_merge_cells(start_row, end_row, start_col, end_col):
        try:
            # 验证合并范围的有效性
            if start_row > end_row or start_col > end_col:
                return False
            if start_row < 1 or start_col < 1:
                return False
            if end_row > worksheet.max_row or end_col > worksheet.max_column:
                return False
            
            # 检查是否与现有合并区域重叠
            for merged_range in get_merged_ranges():
                if (max(start_row, merged_range.min_row) <= min(end_row, merged_range.max_row) and
                    max(start_col, merged_range.min_col) <= min(end_col, merged_range.max_col)):
                    return False
            
            worksheet.merge_cells(
                start_row=start_row,
                end_row=end_row,
                start_column=start_col,
                end_column=end_col
            )
            return True
        except Exception:
            return False
    
    def is_empty_cell(cell):
        try:
            if cell is None:
                return True
            value = cell.value
            if value is None:
                return True
            if isinstance(value, str):
                return not value.strip()
            return False
        except Exception:
            return False
    
    try:
        merged_ranges = get_merged_ranges()
        
        # 第一步：优先处理第五行的横向合并，并检查是否可以向下合并到第七行
        if header_row and header_row + 1 <= end_row:
            fifth_row = header_row + 1
            col = start_col
            while col <= end_col:
                if is_cell_in_merged_range(fifth_row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= fifth_row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        col = merge_range.max_col + 1
                    continue
                
                current_cell = worksheet.cell(row=fifth_row, column=col)
                if not is_empty_cell(current_cell):
                    # 向右查找可以合并的空单元格，直到遇到非空单元格
                    merge_end_col = col
                    for next_col in range(col + 1, end_col + 1):
                        next_cell = worksheet.cell(row=fifth_row, column=next_col)
                        if (is_empty_cell(next_cell) and 
                            not is_cell_in_merged_range(fifth_row, next_col, merged_ranges)):
                            merge_end_col = next_col
                        else:
                            break
                    
                    if merge_end_col > col:
                        # 先进行横向合并
                        if try_merge_cells(fifth_row, fifth_row, col, merge_end_col):
                            merged_ranges = get_merged_ranges()
                            
                            # 检查第六行和第七行的对应列是否为空
                            can_merge_down = True
                            max_merge_row = min(fifth_row + 2, end_row)  # 最多合并到第七行
                            
                            # 检查是否可以向下合并
                            for check_row in range(fifth_row + 1, max_merge_row + 1):
                                for check_col in range(col, merge_end_col + 1):
                                    check_cell = worksheet.cell(row=check_row, column=check_col)
                                    if not is_empty_cell(check_cell) or is_cell_in_merged_range(check_row, check_col, merged_ranges):
                                        can_merge_down = False
                                        break
                                if not can_merge_down:
                                    break
                            
                            # 如果可以向下合并，则合并到第七行（或最大可能行）
                            if can_merge_down and max_merge_row > fifth_row:
                                # 先取消之前的横向合并
                                worksheet.unmerge_cells(
                                    start_row=fifth_row,
                                    end_row=fifth_row,
                                    start_column=col,
                                    end_column=merge_end_col
                                )
                                # 进行完整的合并（横向+纵向）
                                if try_merge_cells(fifth_row, max_merge_row, col, merge_end_col):
                                    merged_ranges = get_merged_ranges()
                    
                    col = merge_end_col + 1
                else:
                    col += 1
        
        # 第二步：处理A-C列的横向合并（跳过第五行）
        text_col_end = min(3, end_col)  # A-C列（如果存在）
        for row in range(start_row, end_row + 1):
            if header_row and row == header_row + 1:  # 跳过第五行
                continue
            
            col = start_col
            while col <= text_col_end:
                if is_cell_in_merged_range(row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        col = merge_range.max_col + 1
                    continue
                
                current_cell = worksheet.cell(row=row, column=col)
                if not is_empty_cell(current_cell):
                    # 向右查找可以合并的空单元格
                    merge_end_col = col
                    for next_col in range(col + 1, text_col_end + 1):
                        next_cell = worksheet.cell(row=row, column=next_col)
                        if (is_empty_cell(next_cell) and 
                            not is_cell_in_merged_range(row, next_col, merged_ranges)):
                            merge_end_col = next_col
                        else:
                            break
                    
                    if merge_end_col > col:
                        if try_merge_cells(row, row, col, merge_end_col):
                            merged_ranges = get_merged_ranges()
                    col = merge_end_col + 1
                else:
                    col += 1
        
        # 第三步：处理A-C列的纵向合并（包括第五行）
        for col in range(1, text_col_end + 1):  # 仅处理A-C列
            row = start_row
            while row <= end_row:
                if is_cell_in_merged_range(row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        row = merge_range.max_row + 1
                    continue
                
                current_cell = worksheet.cell(row=row, column=col)
                if not is_empty_cell(current_cell):
                    # 获取当前单元格所在的横向合并范围
                    current_merge = next((r for r in merged_ranges 
                                        if r.min_row == row and r.min_col <= col <= r.max_col), None)
                    
                    merge_start_col = current_merge.min_col if current_merge else col
                    merge_end_col = current_merge.max_col if current_merge else col
                    
                    # 向下查找可以合并的空行
                    merge_end_row = row
                    for next_row in range(row + 1, end_row + 1):
                        # 检查下一行的整个范围是否为空
                        is_range_empty = True
                        for check_col in range(merge_start_col, merge_end_col + 1):
                            next_cell = worksheet.cell(row=next_row, column=check_col)
                            if not is_empty_cell(next_cell) or is_cell_in_merged_range(next_row, check_col, merged_ranges):
                                is_range_empty = False
                                break
                        
                        if is_range_empty:
                            merge_end_row = next_row
                        else:
                            break
                    
                    if merge_end_row > row:
                        if current_merge:
                            # 先取消横向合并
                            worksheet.unmerge_cells(
                                start_row=row,
                                end_row=row,
                                start_column=merge_start_col,
                                end_column=merge_end_col
                            )
                        # 进行完整的合并
                        if try_merge_cells(row, merge_end_row, merge_start_col, merge_end_col):
                            merged_ranges = get_merged_ranges()
                    row = merge_end_row + 1
                else:
                    row += 1
        
        # 第四步：处理其他列的纵向合并（D列及之后的列）
        for col in range(text_col_end + 1, end_col + 1):
            row = start_row
            while row <= end_row:
                if is_cell_in_merged_range(row, col, merged_ranges):
                    merge_range = next((r for r in merged_ranges 
                                      if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
                    if merge_range:
                        row = merge_range.max_row + 1
                    continue
                
                current_cell = worksheet.cell(row=row, column=col)
                if not is_empty_cell(current_cell):
                    # 向下查找可以合并的空单元格
                    merge_end_row = row
                    for next_row in range(row + 1, end_row + 1):
                        next_cell = worksheet.cell(row=next_row, column=col)
                        if (is_empty_cell(next_cell) and 
                            not is_cell_in_merged_range(next_row, col, merged_ranges)):
                            merge_end_row = next_row
                        else:
                            break
                    
                    if merge_end_row > row:
                        if try_merge_cells(row, merge_end_row, col, col):
                            merged_ranges = get_merged_ranges()
                    row = merge_end_row + 1
                else:
                    row += 1
                    
    except Exception as e:
        print(f"合并单元格时出错: {str(e)}")
        # 发生错误时不影响程序继续运行

    """
    获取鼠标位置的辅助函数
    使用方法：运行此函数，然后将鼠标移动到目标位置，按Ctrl+C结束
    """
    try:
        print("请将鼠标移动到目标位置...")
        print("按Ctrl+C停止并显示坐标")
        while True:
            x, y = pyautogui.position()
            position_str = f'当前鼠标位置: X: {x} Y: {y}'
            print(position_str, end='\r')
            time.sleep(0.1)
    except KeyboardInterrupt:
        print("\n已记录最后位置")
        return x, y

    """
    使用自动化方式将网页中的图片复制到Excel
    Args:
        driver: WebDriver对象
        excel_path: Excel文件路径
        page_numbers: 需要处理的页码列表
    """
    excel = None
    workbook = None
    max_retries = 3  # 最大重试次数
    
    try:
        # 转换为绝对路径
        excel_abs_path = os.path.abspath(excel_path)
        print(f"Excel路径: {excel_abs_path}")
        
        # 检查文件是否存在
        if not os.path.exists(excel_abs_path):
            print(f"Excel文件不存在: {excel_abs_path}")
            return
        
        # 创建临时文件夹（如果不存在）
        temp_dir = os.path.join('downloads', 'temp')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            print(f"创建临时目录: {temp_dir}")
        
        # 打开Excel文件（带重试）
        for attempt in range(max_retries):
            try:
                print(f"尝试打开Excel... (尝试 {attempt + 1}/{max_retries})")
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                excel.DisplayAlerts = False
                workbook = excel.Workbooks.Open(excel_abs_path)
                print("Excel打开成功")
                break
            except Exception as e:
                print(f"打开Excel失败 (尝试 {attempt + 1}): {str(e)}")
                if attempt == max_retries - 1:
                    raise
                if excel:
                    try:
                        excel.Quit()
                    except:
                        pass
                time.sleep(2)  # 等待一段时间后重试
        
        # 设置pyautogui安全设置
        pyautogui.FAILSAFE = True
        pyautogui.PAUSE = 1.5
        
        for page_num in page_numbers:
            try:
                print(f"\n处理第 {page_num} 页...")
                
                # 询问是否需要重新获取坐标
                response = input(f"\n是否需要重新获取第 {page_num} 页的图片坐标？(y/n): ")
                if response.lower() == 'y':
                    print("\n请获取图片左上角坐标:")
                    start_x, start_y = get_mouse_position()
                    print(f"\n左上角坐标: X: {start_x} Y: {start_y}")
                    
                    print("\n请获取图片右下角坐标:")
                    end_x, end_y = get_mouse_position()
                    print(f"\n右下角坐标: X: {end_x} Y: {end_y}")
                else:
                    # 使用上一次的坐标
                    if 'start_x' not in locals():
                        print("没有保存的坐标，请至少获取一次坐标")
                        continue
                
                print("\n请切换到PDF网页，3秒后将自动截图...")
                time.sleep(3)
                
                # 计算区域大小
                width = end_x - start_x
                height = end_y - start_y
                
                # 保存截图到临时文件（带重试）
                temp_image = os.path.join(temp_dir, f'temp_image_{page_num}.png')
                screenshot_success = False
                
                for attempt in range(max_retries):
                    try:
                        print(f"尝试截图... (尝试 {attempt + 1}/{max_retries})")
                        screenshot = pyautogui.screenshot(region=(start_x, start_y, width, height))
                        screenshot.save(temp_image)
                        
                        if os.path.exists(temp_image) and os.path.getsize(temp_image) > 0:
                            screenshot_success = True
                            print("截图成功")
                            break
                    except Exception as e:
                        print(f"截图失败 (尝试 {attempt + 1}): {str(e)}")
                        if attempt == max_retries - 1:
                            print("截图失败，跳过当前页面")
                            continue
                        time.sleep(2)
                
                if not screenshot_success:
                    continue
                
                # 插入图片到Excel（带重试）
                for attempt in range(max_retries):
                    try:
                        print(f"尝试插入图片... (尝试 {attempt + 1}/{max_retries})")
                        
                        # 激活工作表
                        worksheet = workbook.Worksheets(f'第{page_num}页')
                        worksheet.Activate()
                        
                        # 选择插入位置
                        paste_row = 6
                        cell = worksheet.Cells(paste_row, 2)
                        cell.Select()
                        
                        # 获取插入位置
                        left = cell.Left
                        top = cell.Top
                        
                        # 插入图片
                        shape = worksheet.Shapes.AddPicture(
                            os.path.abspath(temp_image),
                            LinkToFile=False,
                            SaveWithDocument=True,
                            Left=left,
                            Top=top,
                            Width=600,
                            Height=400
                        )
                        
                        # 调整行高
                        worksheet.Rows(paste_row).RowHeight = 400
                        
                        print(f"第 {page_num} 页图片插入成功")
                        break
                        
                    except Exception as e:
                        print(f"插入图片失败 (尝试 {attempt + 1}): {str(e)}")
                        if attempt == max_retries - 1:
                            print("插入图片失败，跳过当前页面")
                        time.sleep(2)
                
            except Exception as e:
                print(f"处理第 {page_num} 页时出错: {str(e)}")
                continue
        
        # 保存Excel文件（带重试）
        for attempt in range(max_retries):
            try:
                print(f"\n尝试保存Excel... (尝试 {attempt + 1}/{max_retries})")
                workbook.Save()
                print(f"Excel文件已保存到: {excel_abs_path}")
                break
            except Exception as e:
                print(f"保存Excel失败 (尝试 {attempt + 1}): {str(e)}")
                if attempt == max_retries - 1:
                    print("保存Excel失败")
                time.sleep(2)
        
        # 询问是否关闭Excel
        response = input("\n所有图片处理完成，是否关闭Excel？(y/n): ")
        if response.lower() == 'y':
            try:
                print("关闭Excel...")
                if workbook:
                    workbook.Close()
                if excel:
                    excel.Quit()
            except Exception as e:
                print(f"关闭Excel时出错: {str(e)}")
        else:
            print("Excel保持打开状态")
        
    except Exception as e:
        print(f"处理过程出错: {str(e)}")
        import traceback
        traceback.print_exc()
        
    finally:
        # 清理临时文件
        try:
            if os.path.exists(temp_dir):
                import shutil
                shutil.rmtree(temp_dir)
                print("临时目录已清理")
        except Exception as e:
            print(f"清理临时目录失败: {str(e)}")
        
        # 确保Excel正确关闭
        if 'response' in locals() and response.lower() == 'y':
            try:
                if workbook:
                    workbook.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
            except:
                pass

    """获取云南省电价项目列表"""
    try:
        print("等待电价项目列表加载...")
        price_items = []
        
        # 等待项目列表加载
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'list-item')]"))
        )
        
        # 获取所有项目元素
        items = driver.find_elements(By.XPATH, "//div[contains(@class, 'list-item')]")
        
        for item in items:
            try:
                # 获取项目标题
                title = item.find_element(By.XPATH, ".//div[contains(@class, 'esp')]").text.strip()
                # 获取日期
                date = item.find_element(By.XPATH, ".//div[contains(@class, 'timeLine')]").text.strip()
                
                if title and '电价' in title:  # 只添加包含"电价"的项目
                    link_element = item.find_element(By.XPATH, ".//div[contains(@class, 'link')]")
                    price_items.append({
                        "text": f"{title} ({date})",
                        "element": link_element
                    })
                    print(f"找到电价项目: {title}")
            except Exception as e:
                print(f"处理项目元素时出错: {str(e)}")
                continue
        
        return price_items
    except Exception as e:
        print(f"获取电价项目列表失败: {str(e)}")
        return []


    """调整云南省表格格式"""
    try:
        # 检查是否是DataFrame
        if not isinstance(data, pd.DataFrame):
            return data
            
        # 重命名列（如果需要）
        column_mapping = {
            '用户类别': '电压等级',
            '基本电价': '基本电费',
            '电度电价': '电度电费'
        }
        data = data.rename(columns=lambda x: column_mapping.get(x, x))
        
        # 调整列顺序（根据云南省的要求）
        desired_columns = ['电压等级', '基本电费', '电度电费']
        existing_columns = [col for col in desired_columns if col in data.columns]
        other_columns = [col for col in data.columns if col not in desired_columns]
        data = data[existing_columns + other_columns]
        
        return data
    except Exception as e:
        print(f"调整云南省表格格式失败: {str(e)}")
        return data
def apply_province_specific_format(worksheet, is_yunnan=False):
    """根据省份应用特定的格式"""
    if is_yunnan:
        apply_yunnan_styles(worksheet)
        apply_yunnan_cell_format(worksheet)
    else:
        apply_cell_format(worksheet)  # 通用格式

    """应用云南省特有的样式"""
    try:
        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置对齐方式
        center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        left_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )
        
        # 设置字体
        regular_font = Font(name='宋体', size=11)
        title_font = Font(name='宋体', size=16, bold=True)
        subtitle_font = Font(name='宋体', size=14, bold=True)
        
        # 获取工作表名称
        sheet_name = worksheet.title
        page_number = int(sheet_name.replace('第', '').replace('页', ''))
        
        # 应用基础格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.font = regular_font
                cell.alignment = left_alignment  # 默认使用左对齐
        
        if page_number <= 3:  # 前三页
            # 获取所有合并单元格的范围
            merged_ranges = worksheet.merged_cells.ranges
            
            # 对合并的单元格应用居中对齐
            for merged_range in merged_ranges:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = center_alignment
            
            # 设置标题行格式（第1行）
            for cell in worksheet[1]:
                cell.font = title_font
            
            # 设置副标题行格式（第2行）
            for cell in worksheet[2]:
                cell.font = title_font  # 使用与第1行相同的字体
            
            # 设置单位行格式（第3行）
            for cell in worksheet[3]:
                cell.font = subtitle_font
            
        else:  # 后三页
            # 前三行使用左对齐、自动换行
            for row in range(1, 4):
                for cell in worksheet[row]:
                    cell.alignment = left_alignment
                    if row == 1:
                        cell.font = title_font
                    elif row == 2:
                        cell.font = title_font
                    else:  # row == 3
                        cell.font = subtitle_font
                
                # 合并单元格（A到J列）
                if not any(worksheet.merged_cells.ranges, key=lambda r: r.min_row == row):
                    worksheet.merge_cells(f'A{row}:J{row}')
            
            # 其余行使用与前三行一致的格式
            for row in range(4, worksheet.max_row + 1):
                for cell in worksheet[row]:
                    cell.alignment = left_alignment
        
        # 设置行高
        worksheet.row_dimensions[1].height = 35  # 标题行
        worksheet.row_dimensions[2].height = 35  # 副标题行（与第1行相同）
        worksheet.row_dimensions[3].height = 25  # 单位行
        
        if page_number <= 3:  # 前三页
            worksheet.row_dimensions[4].height = 30  # 普通行高
            worksheet.row_dimensions[5].height = 30
        else:  # 后三页
            worksheet.row_dimensions[i].height = 47  # 修改为47
            
    except Exception as e:
        print(f"应用云南省样式失败: {str(e)}")
        import traceback
        traceback.print_exc()

    """云南省特有的单元格合并逻辑"""
    try:
        # 设置边框样式
        thin_side = Side(style='thin')
        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )
        
        # 获取特殊列的索引
        special_merge_columns = ['电压等级', '基本电费', '电度电费']
        special_col_indices = []
        
        if header_row:
            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet.cell(row=header_row, column=col)
                if header_cell.value in special_merge_columns:
                    special_col_indices.append(col)
        
        # 对特殊列应用合并规则
        for col in special_col_indices:
            current_value = None
            merge_start = None
            
            for row in range(start_row, end_row + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value == current_value and cell_value is not None:
                    continue
                elif merge_start and merge_start < row - 1:
                    # 合并前面的相同值单元格
                    worksheet.merge_cells(
                        start_row=merge_start,
                        end_row=row - 1,
                        start_column=col,
                        end_column=col
                    )
                    # 应用边框
                    for r in range(merge_start, row):
                        worksheet.cell(row=r, column=col).border = thin_border
                
                current_value = cell_value
                merge_start = row if cell_value is not None else None
            
            # 处理最后一组相同值
            if merge_start and merge_start < end_row:
                worksheet.merge_cells(
                    start_row=merge_start,
                    end_row=end_row,
                    start_column=col,
                    end_column=col
                )
                # 应用边框
                for r in range(merge_start, end_row + 1):
                    worksheet.cell(row=r, column=col).border = thin_border
                
    except Exception as e:
        print(f"云南省单元格合并失败: {str(e)}")


    """添加云南省特有的注释格式"""
    try:
        if not notes:
            return
            
        # 添加"注释内容"标题行
        title_row = start_row + 1
        title_cell = worksheet.cell(row=title_row, column=1, value="注释说明")
        
        # 合并标题行
        title_range = f'A{title_row}:{openpyxl.utils.get_column_letter(max_col)}{title_row}'
        worksheet.merge_cells(title_range)
        
        # 设置标题样式
        title_font = Font(name='宋体', size=12, bold=True)
        title_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        title_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 应用标题样式
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        
        # 设置注释样式
        note_font = Font(name='宋体', size=11)
        note_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 写入注释内容
        for i, note in enumerate(notes, 1):
            note_row = title_row + i
            note_cell = worksheet.cell(row=note_row, column=1, value=note)
            
            # 合并注释行
            note_range = f'A{note_row}:{openpyxl.utils.get_column_letter(max_col)}{note_row}'
            worksheet.merge_cells(note_range)
            
            # 应用注释样式
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # 设置行高
            worksheet.row_dimensions[note_row].height = 30
            
    except Exception as e:
        print(f"添加云南省注释失败: {str(e)}")


    """应用云南省特有的单元格格式"""
    try:
        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置对齐方式
        center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        # 设置字体
        regular_font = Font(name='宋体', size=11)
        
        # 应用格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = regular_font
        
        # 设置标题行特殊格式
        title_font = Font(name='宋体', size=16, bold=True)
        subtitle_font = Font(name='宋体', size=14, bold=True)
        
        for cell in worksheet[1]:
            cell.font = title_font
        for cell in worksheet[2:4]:
            for c in cell:
                c.font = subtitle_font
        
        # 设置表头行特殊格式
        header_font = Font(name='宋体', size=12, bold=True)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        for cell in worksheet[4]:
            cell.font = header_font
            cell.fill = header_fill
            
    except Exception as e:
        print(f"应用云南省单元格格式失败: {str(e)}")
def main():
    """主函数"""
    print("=== 开始运行自动化程序 ===")
    
    driver = None
    try:
        # 初始化驱动
        driver = setup_driver()
        if not driver:
            return
        
        # 访问网站
        print("正在访问网站...")
        driver.get("https://www.95598.cn/osgweb/index")
        time.sleep(3)
        
        while True:  # 主循环
            # 第一步：选择省份
            if not handle_province_selection(driver):
                continue
            
            # 第二步：导航到电价标准页面
            if not handle_price_page_navigation(driver):
                driver.get("https://www.95598.cn/osgweb/index")
                time.sleep(2)
                continue
            
            while True:  # 电价标准页面循环
                # 第三步：选择城市
                if not handle_city_selection(driver):
                    driver.get("https://www.95598.cn/osgweb/index")
                    time.sleep(2)
                    break
                
                # 第四步：选择区县
                if not handle_district_selection(driver):
                    driver.get("https://www.95598.cn/osgweb/index")
                    time.sleep(2)
                    break
                
                # 第五步：选择电价标准项目
                result = handle_project_selection(driver)
                if result == "back_to_price_page":
                    continue  # 继续电价标准页面循环
                elif not result:
                    driver.get("https://www.95598.cn/osgweb/index")
                    time.sleep(2)
                    break
                
                # 第六步：选择具体项目
                result = handle_subproject_selection(driver)
                if result == "exit":
                    return
                elif not result:
                    driver.get("https://www.95598.cn/osgweb/index")
                    time.sleep(2)
                    break
                
                # 处理子项目导航
                result = handle_subproject_navigation(driver, result)
                if result == "exit":
                    return
                elif result == "back_to_price_page":
                    continue  # 继续电价标准页面循环
                elif not result:
                    driver.get("https://www.95598.cn/osgweb/index")
                    time.sleep(2)
                    break
            
            # 云南省特殊处理
            if selected_province['text'] == '云南省':
                try:
                    print("\n检测到云南省，执行特殊处理流程...")
                    
                    # 点击营商环境
                    print("等待信息公开加载...")
                    business_env = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//li[@data-v-4781518a and @class='li_lis']//span[@id='right-class' and contains(text(), '信息公开')]"))
                    )
                    
                    if not wait_and_click_element(driver, business_env):
                        print("点击信息公开失败")
                        continue
                    
                    # 点击电价标准
                    print("等待电价标准加载...")
                    info_disclosure = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'submenu')]//a[text()='电价标准']"))
                    )
                    
                    if not wait_and_click_element(driver, info_disclosure):
                        print("点击信息公开失败")
                        continue
                    
                    
                    time.sleep(2)  # 等待页面加载
                    
                    # 获取电价项目列表
                    price_items = get_yunnan_price_items(driver)
                    if not price_items:
                        print("未找到电价项目列表")
                        continue
                    
                    # 选择项目
                    selected_item = display_menu(price_items, "请选择电价项目：")
                    if selected_item is None:
                        continue
                    
                    print(f"\n正在处理: {selected_item['text']}")
                    if not wait_and_click_element(driver, selected_item['element']):
                        print("点击项目失败")
                        continue
                    
                    time.sleep(2)  # 等待页面加载
                    
                    # 检查是否有PDF并处理
                    excel_file = check_and_process_content(driver, selected_item['text'], is_yunnan=True)
                    if excel_file:
                        print(f"\n处理完成，数据已保存到: {excel_file}")
                    else:
                        print("\n处理失败")
                
                except Exception as e:
                    print(f"云南省特殊处理失败: {str(e)}")
                    continue
            else:
                # 其他省份的原有处理逻辑# 获取总页数
                total_pages = get_page_info(driver)
                if total_pages == 0:
                    print("获取页码信息失败")
                    continue
                
                while True:  # 页码选择循环
                    print("\n=== 可选页码 ===")
                    for page in range(1, total_pages + 1):
                        print(f"{page}. 第 {page} 页")
                    print("0. 返回上一步")
                    
                    try:
                        page_choice = int(input("\n请输入页码: "))
                        if page_choice == 0:
                            break
                        
                        if 1 <= page_choice <= total_pages:
                            announcements = get_page_announcements(driver, page_choice)
                            if not announcements:
                                print(f"第 {page_choice} 页没有找到公告")
                                continue
                            
                            print(f"\n=== 第 {page_choice} 页的公告列表 ===")
                            for i, item in enumerate(announcements, 1):
                                print(f"{i}. {item['text']}")
                            print("0. 返回上一步")
                            
                            try:
                                item_choice = int(input(f"\n请输入编号(0-{len(announcements)}): "))
                                if item_choice == 0:
                                    continue
                                
                                if 1 <= item_choice <= len(announcements):
                                    selected_announcement = announcements[item_choice-1]
                                    print(f"\n您选择了: {selected_announcement['text']}")
                                    
                                    print(f"\n正在处理: {selected_announcement['text']}")
                                    if not wait_and_click_element(driver, selected_announcement['element']):
                                        print("点击公告失败")
                                        continue
                                    
                                    time.sleep(2)  # 等待页面加载
                                    
                                    # 检查是否有PDF并处理
                                    excel_file = check_and_process_content(driver, selected_announcement['text'], is_yunnan=False)
                                    if excel_file:
                                        print(f"\n处理完成，数据已保存到: {excel_file}")
                                    else:
                                        print("\n处理失败")
                                    
                                    while True:  # 操作选择循环
                                        print("\n请选择操作：")
                                        print("1. 继续浏览")
                                        print("2. 返回上一步")
                                        print("3. 退出程序")
                                        
                                        try:
                                            choice = int(input("\n请输入编号(1-3): "))
                                            if choice == 1:
                                                break
                                            elif choice == 2:
                                                break
                                            elif choice == 3:
                                                return
                                            else:
                                                print("无效的选择，请输入1-3之间的数字")
                                        except ValueError:
                                            print("请输入有效的数字")
                                else:
                                    print(f"无效的选择，请输入0-{len(announcements)}之间的数字")
                            except ValueError:
                                print("请输入有效的数字")
                        else:
                            print(f"无效的页码，请输入0-{total_pages}之间的数字")
                    except ValueError:
                        print("请输入有效的页码")
            
            # 操作选择循环
            while True:
                print("\n请选择操作：")
                print("1. 继续浏览")
                print("2. 返回上一步")
                print("3. 退出程序")
                
                try:
                    choice = int(input("\n请输入编号(1-3): "))
                    if choice == 1:
                        break
                    elif choice == 2:
                        break
                    elif choice == 3:
                        return
                    else:
                        print("无效的选择，请输入1-3之间的数字")
                except ValueError:
                    print("请输入有效的数字")
        
    except Exception as e:
        print(f"\n程序运行出错: {str(e)}")
    finally:
        if driver:
            driver.quit()
def handle_province_selection(driver):
    """处理省份选择"""
    print("\n第一步：选择省份")
    
    # 点击地区选择器
    region_selectors = [
        "//div[@id='city_select']//a[contains(@class, 'current')]",
        "//div[contains(@class, 'region')]//a[contains(@class, 'current')]",
        "//a[contains(@class, 'current fsize16')]"
    ]
    
    region_element = find_element_with_retry(driver, region_selectors, timeout=3)
    if not region_element or not wait_and_click_element(driver, region_element):
        print("点击地区选择器失败")
        return False
    
    time.sleep(0.5)
    
    # 获取省份列表
    province_selectors = ["//a[contains(@class, 'f66 fsize14')]"]  # 简化选择器
    
    provinces = []
    for selector in province_selectors:
        elements = get_visible_elements(driver, selector)
        for element in elements:
            text = extract_element_text(element)
            if text and text != '省份' and text not in [p['text'] for p in provinces]:
                provinces.append({"text": text, "element": element})
    
    if not provinces:
        print("未找到任何省份")
        return False
    
    # 用户选择省份
    selected_province = display_menu(provinces, "请选择省份：")
    if not selected_province:
        return False
    
    # 点击选中的省份
    return wait_and_click_element(driver, selected_province['element'])
def handle_price_page_navigation(driver):
    """处理电价标准页面导航"""
    print("\n第二步：导航到电价标准页面")
    time.sleep(0.5)
    
    # 点击信息公开
    info_selectors = ["//span[@id='right-class' and contains(text(), '信息公开')]"]
    info_btn = find_element_with_retry(driver, info_selectors, timeout=3)
    if not info_btn or not wait_and_click_element(driver, info_btn):
        print("点击信息公开失败")
        return False
    
    time.sleep(0.5)
    
    # 点击电价标准
    price_selectors = ["//a[text()='电价标准']"]
    price_btn = find_element_with_retry(driver, price_selectors, timeout=3)
    if not price_btn or not wait_and_click_element(driver, price_btn):
        print("点击电价标准失败")
        return False
    
    return True
def handle_project_navigation(driver, project):
    """处理项目导航"""
    try:
        # 确保元素在视图中
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", project['element'])
        time.sleep(1)
        
        # 检查是否已经展开
        expand_icons = project['element'].find_elements(By.CSS_SELECTOR, ".el-table__expand-icon")
        if expand_icons:
            expand_icon = expand_icons[0]
            is_expanded = "expanded" in expand_icon.get_attribute("class")
            
            if not is_expanded:
                # 尝试点击展开箭头
                try:
                    expand_icon.click()
                except:
                    try:
                        driver.execute_script("arguments[0].click();", expand_icon)
                    except:
                        ActionChains(driver).move_to_element(expand_icon).click().perform()
                print("已点击展开箭头")
            else:
                print("项目已经是展开状态")
            
            time.sleep(1)
            return True
            
        else:
            print("未找到展开箭头")
            return False
            
    except Exception as e:
        print(f"项目导航失败: {str(e)}")
        return False

def handle_subproject_navigation(driver, subproject):
    """处理子项目导航"""
    try:
        # 确保元素在视图中并可点击
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", subproject['element'])
        time.sleep(1)
        
        # 尝试移除可能的遮挡元素
        driver.execute_script("""
            var overlays = document.querySelectorAll('.tab-content, .el-dialog, .el-dialog__wrapper');
            overlays.forEach(function(overlay) {
                overlay.style.pointerEvents = 'none';
            });
        """)
        
        if not wait_and_click_element(driver, subproject['element']):
            # 如果点击失败，尝试直接获取链接地址并导航
            try:
                href = subproject['element'].get_attribute('href')
                if href:
                    print(f"\n尝试直接导航到地址: {href}")
                    driver.get(href)
                else:
                    return False
            except:
                return False
        
        print("\n已进入选择的页面...")
        time.sleep(1)
        
        while True:
            print("\n当前操作选项：")
            print("1. 继续浏览当前页面")
            print("2. 返回重新选择城市")
            print("3. 退出程序") 
            print("4. 保存当前页面为PDF")  # 修改选项描述
            
            action = input("\n请选择操作 (1/2/3/4): ").strip()
            
            if action == "1":
                print("\n当前页面浏览选项：")
                print("1. 向下滚动")
                print("2. 向上滚动")
                print("3. 回到顶部")
                print("4. 回到底部")
                print("5. 返回上一级菜单")
                
                browse_action = input("\n请选择浏览操作 (1/2/3/4/5): ").strip()
                
                if browse_action == "1":
                    driver.execute_script("window.scrollBy(0, 500);")
                elif browse_action == "2":
                    driver.execute_script("window.scrollBy(0, -500);")
                elif browse_action == "3":
                    driver.execute_script("window.scrollTo(0, 0);")
                elif browse_action == "4":
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                elif browse_action == "5":
                    continue
                else:
                    print("无效的选择，请重试")
                continue
                
            elif action == "2":
                print("\n正在返回到电价标准页面...")
                # 直接重新加载电价标准页面
                info_btn = find_element_with_retry(driver, ["//span[@id='right-class' and contains(text(), '信息公开')]"])
                if info_btn and wait_and_click_element(driver, info_btn):
                    time.sleep(1)
                    price_btn = find_element_with_retry(driver, ["//a[text()='电价标准']"])
                    if price_btn and wait_and_click_element(driver, price_btn):
                        time.sleep(1)
                        return "back_to_price_page"
                
                # 如果直接导航失败，则重新加载首页
                print("正在重新加载页面...")
                driver.get("https://www.95598.cn/osgweb/index")
                time.sleep(2)
                if handle_province_selection(driver) and handle_price_page_navigation(driver):
                    return "back_to_price_page"
                return False
                
            elif action == "3":
                confirm = input("\n确定要退出程序吗？(y/n): ").strip().lower()
                if confirm == 'y':
                    print("\n正在退出程序...")
                    return "exit"
                continue
            
            elif action == "4":
                try:
                    print("\n正在准备保存PDF...")
                    # 创建downloads目录（如果不存在）
                    if not os.path.exists('downloads'):
                        os.makedirs('downloads')
                    
                    # 获取项目名称并清理非法字符
                    project_name = subproject['text']
                    # 清理Windows文件名中的非法字符
                    project_name = re.sub(r'[<>:"/\\|?*]', '_', project_name)
                    
                    # 生成基础文件名
                    base_filename = project_name
                    
                    # 检查是否存在同名文件，并添加版本号
                    version = 1
                    while True:
                        filename = os.path.join('downloads', f'{base_filename}_V{version}.pdf')
                        if not os.path.exists(filename):
                            break
                        version += 1
                    
                    print(f"将保存为: {os.path.basename(filename)}")
                    
                    # 获取当前页面URL
                    current_url = driver.current_url
                    print(f"原始URL: {current_url}")  # 调试输出
                    
                    # 检查是否有新窗口打开
                    all_handles = driver.window_handles
                    if len(all_handles) > 1:
                        # 切换到最新打开的窗口
                        driver.switch_to.window(all_handles[-1])
                        current_url = driver.current_url
                        print(f"新窗口URL: {current_url}")
                    else:
                        # 检查是否有iframe
                        try:
                            iframes = driver.find_elements(By.TAG_NAME, "iframe")
                            for iframe in iframes:
                                src = iframe.get_attribute('src')
                                if src and ('.pdf' in src.lower() or 'omg-static' in src.lower()):
                                    current_url = src
                                    print(f"找到PDF iframe URL: {current_url}")
                                    break
                        except:
                            pass
                    
                    # 如果URL包含PDF或omg-static（不区分大小写）
                    if '.pdf' in current_url.lower() or 'omg-static' in current_url.lower():
                        print(f"检测到PDF文件URL: {current_url}")
                        headers = {
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                            'Accept': 'application/pdf,*/*',
                            'Accept-Encoding': 'gzip, deflate, br',
                            'Connection': 'keep-alive',
                            'Referer': 'https://95598.cn/omg-static//omg-static/'
                        }
                        
                        try:
                            # 禁用SSL验证警告
                            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                            
                            # 使用session来处理请求
                            session = requests.Session()
                            session.verify = False  # 禁用SSL验证
                            
                            # 直接下载PDF
                            print("开始下载PDF...")  # 添加调试输出
                            response = session.get(current_url, headers=headers, timeout=30)
                            print(f"响应状态码: {response.status_code}")  # 添加调试输出
                            
                            if response.status_code == 200:
                                # 检查内容类型和文件大小
                                content_length = len(response.content)
                                print(f"下载内容大小: {content_length} bytes")  # 添加调试输出
                                
                                if content_length == 0:
                                    print("下载的内容为空")
                                    return None
                                
                                # 保存文件
                                with open(filename, 'wb') as f:
                                    f.write(response.content)
                                
                                # 验证保存的文件
                                if os.path.exists(filename) and os.path.getsize(filename) > 0:
                                    print(f"\nPDF文件已成功保存到: {filename}")
                                    # 删除return语句，继续循环
                                    continue  # 返回到主菜单
                                else:
                                    print("保存的文件为空或不存在")
                                    continue  # 返回到主菜单
                            else:
                                print(f"\n下载失败，状态码: {response.status_code}")
                                continue  # 返回到主菜单
                        except Exception as e:
                            print(f"下载PDF时发生错误: {str(e)}")
                            continue  # 返回到主菜单
                    else:
                        print("\n当前页面不是PDF文件，尝试查找PDF链接...")
                        # 查找页面中的PDF链接
                        pdf_links = driver.find_elements(By.XPATH, "//a[contains(@href, '.pdf')]")
                        if pdf_links:
                            pdf_url = pdf_links[0].get_attribute('href')
                            print(f"找到PDF链接: {pdf_url}")
                            # 下载PDF
                            pdf_response = requests.get(pdf_url, headers=headers, verify=False)
                            if pdf_response.status_code == 200:
                                with open(filename, 'wb') as f:
                                    f.write(pdf_response.content)
                                print(f"\nPDF文件已保存到: {filename}")
                            else:
                                print("\n下载PDF失败")
                        else:
                            print("\n未找到PDF链接")
                        continue  # 返回到主菜单
                except Exception as e:
                    print(f"\n保存PDF失败: {str(e)}")
                continue  # 返回到主菜单
            
            else:
                print("无效的选择，请重试")
        
    except Exception as e:
        print(f"子项目导航失败: {str(e)}")
        return False
    
def handle_city_selection(driver):
    """处理城市选择"""
    print("\n第三步：选择城市")
    time.sleep(0.5)  # 保留短暂等待
    
    cities = get_cities(driver)
    if not cities:
        print("未找到城市列表")
        return False
    
    selected_city = display_menu(cities, "请选择城市：")
    if not selected_city:
        return False
    
    return wait_and_click_element(driver, selected_city['element'])
def handle_district_selection(driver):
    """处理区县选择"""
    print("\n第四步：选择区县")
    time.sleep(0.5)  # 保留短暂等待
    
    districts = get_districts(driver)
    if not districts:
        print("未找到区县列表")
        return False
    
    selected_district = display_menu(districts, "请选择区县：")
    if not selected_district:
        return False
    
    return wait_and_click_element(driver, selected_district['element'])
    
def handle_project_selection(driver):
    """处理项目选择"""
    print("\n第五步：选择电价标准项目")
    
    # 获取项目列表
    projects = get_projects(driver)
    if not projects:
        print("未找到电价标准项目，尝试刷新页面...")
        driver.refresh()
        time.sleep(1)
        projects = get_projects(driver)
        if not projects:
            return False
    
    print("\n当前操作选项：")
    print("0. 返回重新选择城市")
    for i, project in enumerate(projects, 1):
        print(f"{i}. {project['text']}")
    
    try:
        choice = int(input("\n请输入编号: "))
        if choice == 0:
            print("\n正在返回到电价标准页面...")
            # 直接重新加载电价标准页面
            info_btn = find_element_with_retry(driver, ["//span[@id='right-class' and contains(text(), '信息公开')]"])
            if info_btn and wait_and_click_element(driver, info_btn):
                time.sleep(1)
                price_btn = find_element_with_retry(driver, ["//a[text()='电价标准']"])
                if price_btn and wait_and_click_element(driver, price_btn):
                    time.sleep(1)
                    return "back_to_price_page"
            
            # 如果直接导航失败，则重新加载首页
            print("正在重新加载页面...")
            driver.get("https://www.95598.cn/osgweb/index")
            time.sleep(2)
            if handle_province_selection(driver) and handle_price_page_navigation(driver):
                return "back_to_price_page"
            return False
            
        if 1 <= choice <= len(projects):
            selected_project = projects[choice-1]
            print("\n" + "="*50)
            print(f"您选择了: {selected_project['text']}")
            print("="*50 + "\n")
            return handle_project_navigation(driver, selected_project)
        print("无效的选择，请重试")
        return False
    except ValueError:
        print("请输入有效的数字")
        return False
def handle_subproject_selection(driver):
    """处理子项目选择"""
    print("\n第六步：获取子项目列表")
    time.sleep(2)  # 增加等待时间
    
    try:
        # 首先等待展开行出现
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "tr.el-table__expanded-row"))
            )
        except:
            print("\n未找到展开行，尝试其他方式...")
        
        # 使用更精确的选择器
        selectors = [
            # 方式1：通过缩进和占位符定位
            "//div[contains(@class, 'cell')]/span[@class='el-table__indent']/following-sibling::span[@class='el-table__placeholder']/following-sibling::*",
            # 方式2：直接通过cell类和缩进定位
            "//div[contains(@class, 'cell')][.//span[@class='el-table__indent']]",
            # 方式3：通过展开行和缩进结构定位
            "//tr[contains(@class, 'el-table__expanded-row')]//div[contains(@class, 'cell')][.//span[@class='el-table__indent']]"
        ]
        
        cells = []
        for selector in selectors:
            cells = driver.find_elements(By.XPATH, selector)
            if cells:
                print(f"\n使用选择器 '{selector}' 找到 {len(cells)} 个单元格")
                # 打印找到的内容用于调试
                print("\n找到的内容:")
                for i, cell in enumerate(cells, 1):
                    text = cell.text.strip()
                    if text:  # 只打印非空内容
                        print(f"{i}. {text}")
                break
        
        if not cells:
            print("\n未找到展开的内容")
            return False
        
        # 处理每个单元格中的文本
        subprojects = []
        seen_texts = set()
        parent_text = None
        
        # 获取父项目文本用于过滤
        try:
            parent_row = driver.find_element(By.XPATH, "//tr[contains(@class, 'el-table__row')]")
            parent_text = parent_row.text.strip()
            print(f"\n父项目文本: {parent_text}")
        except:
            print("\n无法获取父项目文本")
        
        for cell in cells:
            try:
                text = cell.text.strip()
                if (text and 
                    text not in seen_texts and 
                    not text.startswith("序号") and 
                    not text.startswith("文号") and
                    not text.startswith("发布日期") and
                    not text.startswith("实施日期") and
                    not text.startswith("序") and
                    not text.startswith("号") and
                    not text.isdigit() and  # 排除纯数字
                    "展开" not in text and
                    "收起" not in text and
                    len(text) > 1 and
                    (not parent_text or text != parent_text)):  # 排除与父项目相同的文本
                    seen_texts.add(text)
                    subprojects.append({"text": text, "element": cell})
            except Exception as e:
                print(f"处理单元格内容失败: {str(e)}")
                continue
        
        if subprojects:
            print(f"\n成功获取到 {len(subprojects)} 个子项目")
            print("\n子项目列表:")
            for i, project in enumerate(subprojects, 1):
                print(f"{i}. {project['text']}")
            
            # 添加返回选项
            print("0. 返回上一级")
            
            # 获取用户选择
            try:
                choice = int(input("\n请输入编号 (0-{0}): ".format(len(subprojects))))
                if choice == 0:
                    print("\n返回上一级...")
                    return None
                elif 1 <= choice <= len(subprojects):
                    selected_subproject = subprojects[choice-1]
                    print("\n" + "="*50)
                    print(f"您选择了: {selected_subproject['text']}")
                    print("="*50 + "\n")
                    return selected_subproject
                else:
                    print("\n无效的选择，请重试")
                    return False
            except ValueError:
                print("\n请输入有效的数字")
                return False
        else:
            print("\n未找到任何有效的子项目")
            print("提示：请确保已正确展开主项目，且主项目下包含子项目")
            return False
            
    except Exception as e:
        print(f"\n获取子项目列表失败: {str(e)}")
        print("错误详情:", str(e))
        import traceback
        print(traceback.format_exc())
        return False

if __name__ == "__main__":
    main()
